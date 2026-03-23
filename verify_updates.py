from openpyxl import load_workbook

print("验证更新后的文件：")
print("-" * 50)

# 验证附件3的标题（检查semester是否更新为2026年春季学期）
file_path3 = "期初资料1/2026春季学期_期初资料_任渝/FIN0008B_财经公文写作/财经公文写作-课程组期初教学资料检查情况记录表-任渝.xlsx"
wb3 = load_workbook(file_path3)
ws3 = wb3.active

print("附件3 - 财经公文写作标题：")
print(f"{ws3.cell(row=1, column=1).value}")
print()

# 验证附件1的标题
file_path1 = "期初资料1/2026春季学期_期初资料_任渝/FIN0008B_财经公文写作/智能金融学院-教学大纲审核汇总表-任渝.xlsx"
wb1 = load_workbook(file_path1)
ws1 = wb1.active

print("附件1 - 财经公文写作标题：")
print(f"{ws1.cell(row=1, column=1).value}")
