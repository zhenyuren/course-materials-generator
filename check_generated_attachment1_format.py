from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# 查看生成的附件1文件
wb = load_workbook('智能金融学院_未知教师/FIN3004A_大数据分析基础/数字经济-教学大纲审核汇总表-任渝.xlsx')
ws = wb.active

print('附件1生成文件信息:')
print('=' * 80)
print(f'工作表名称: {ws.title}')
print(f'总行数: {ws.max_row}')
print(f'总列数: {ws.max_column}')

# 检查合并单元格
merged_list = list(ws.merged_cells.ranges)
print(f'合并单元格数量: {len(merged_list)}')
for i, merged_cell in enumerate(merged_list, 1):
    print(f'  {i}. {merged_cell}')

# 检查前5行的内容和格式
print('\n前5行内容:')
for row_idx in range(1, 6):
    row_data = []
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell_value = cell.value
        row_data.append(str(cell_value) if cell_value is not None else 'None')
    print(f'行 {row_idx}: {", ".join(row_data)}')

# 检查第一行的字体和对齐方式
print('\n第一行格式信息:')
cell = ws.cell(row=1, column=1)
font = cell.font
alignment = cell.alignment
print(f'单元格 A1:')
print(f'  字体: {font.name}, 大小: {font.size}, 加粗: {font.bold}')
print(f'  对齐方式: 水平={alignment.horizontal}, 垂直={alignment.vertical}')

# 检查表头的字体和对齐方式
print('\n表头格式信息:')
for col_idx in range(1, 9):
    cell = ws.cell(row=2, column=col_idx)
    font = cell.font
    print(f'单元格 {get_column_letter(col_idx)}2: 字体={font.name}, 大小={font.size}, 加粗={font.bold}')