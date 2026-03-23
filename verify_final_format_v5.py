from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# 读取生成的文件
wb = load_workbook('智能金融学院_未知教师/FIN3004A_大数据分析基础/大数据分析基础-课程组期初教学资料检查情况记录表-任渝.xlsx')
ws = wb.active

print('最终格式验证V5:')
print('=' * 80)

# 检查表末尾端结构位置
print('表末尾端结构位置:')
total_cell = ws.cell(row=19, column=1)
print(f'合计行（第19行）: {total_cell.value}')

materials_cell = ws.cell(row=19, column=5)
print(f'资料份（第19行）: {materials_cell.value}')

# 总体评价意见只占一格（第19行第7列）
evaluation_cell = ws.cell(row=19, column=7)
print(f'总体评价意见（第19行第7列，只占一格）: {evaluation_cell.value}')

# 总体评价意见后面的H到N的格子是合并的，不设置灰色背景
h19_cell = ws.cell(row=19, column=8)
print(f'H19单元格（总体评价意见后面）: {h19_cell.value}')

# 检查合并单元格
print('\n合并单元格:')
merged_list = list(ws.merged_cells.ranges)
for merged_cell in merged_list:
    print(f'  {merged_cell}')

# 检查H19:N19是否合并
print('\nH19:N19合并检查:')
h19n19_merged = 'H19:N19' in [str(merged) for merged in merged_list]
print(f'H19:N19是否合并: {h19n19_merged}')

# 检查总体评价意见单元格G19是否为合并单元格
print('\n总体评价意见单元格检查:')
cell_address = 'G19'
is_merged = any(cell_address in str(merged) for merged in merged_list)
print(f'总体评价意见单元格G19是否为合并单元格: {is_merged}')

# 检查H19单元格的背景色（应该是默认颜色，不是灰色）
print('\nH19单元格背景色检查:')
print(f'H19背景色索引: {h19_cell.fill.fgColor.index}')

# 对比灰色背景的单元格（评价指标列）
print('\n评价指标列背景色（用于对比）:')
metric_cell = ws.cell(row=3, column=9)
print(f'评价指标列背景色索引: {metric_cell.fill.fgColor.index}')

# 判断H19是否有灰色背景
print('\n背景色对比:')
if h19_cell.fill.fgColor.index == '00000000':
    print('✅ H19单元格没有灰色背景（背景色索引: 00000000）')
else:
    print(f'❌ H19单元格有背景色: {h19_cell.fill.fgColor.index}')

if metric_cell.fill.fgColor.index == '00DDDDDD':
    print('✅ 评价指标列有灰色背景（背景色索引: 00DDDDDD）')
else:
    print(f'❌ 评价指标列背景色: {metric_cell.fill.fgColor.index}')

print('\n' + '=' * 80)
print('验证完成！')