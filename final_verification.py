import os
from openpyxl import load_workbook

print('=' * 80)
print('📋 期初资料生成器最终验证')
print('=' * 80)

# 检查输出目录
output_dir = '/Users/dudu/Desktop/智能金融学院/智能体设计/智能金融学院_未知教师'
if not os.path.exists(output_dir):
    print('❌ 输出目录不存在')
    exit(1)

print(f'📁 输出目录: {output_dir}')

# 获取所有课程文件夹
course_folders = [f for f in os.listdir(output_dir) if os.path.isdir(os.path.join(output_dir, f))]
print(f'\n📚 生成的课程文件夹数: {len(course_folders)}')

for folder in course_folders:
    folder_path = os.path.join(output_dir, folder)
    files = os.listdir(folder_path)
    
    print(f'\n{"="*60}')
    print(f'📂 课程文件夹: {folder}')
    print(f'📄 文件数: {len(files)}')
    
    # 检查附件3文件
    attachment3_files = [f for f in files if '附件3' in f or '课程组期初教学资料检查情况记录表' in f]
    if attachment3_files:
        attachment3_file = attachment3_files[0]
        file_path = os.path.join(folder_path, attachment3_file)
        
        print(f'\n📊 附件3文件: {attachment3_file}')
        
        try:
            wb = load_workbook(file_path)
            ws = wb.active
            
            # 检查标题
            title = ws.cell(row=1, column=1).value
            print(f'   标题: {title}')
            
            # 检查数据填充
            course_code = ws.cell(row=5, column=2).value
            course_name = ws.cell(row=5, column=3).value
            department = ws.cell(row=5, column=4).value
            applicable_scope = ws.cell(row=5, column=5).value
            teacher_name = ws.cell(row=5, column=7).value
            course_leader = ws.cell(row=5, column=8).value
            
            print(f'   课程代码: {course_code}')
            print(f'   课程名: {course_name}')
            print(f'   开课单位: {department}')
            print(f'   使用年级/层次/专业: {applicable_scope}')
            print(f'   教师（执笔人）: {teacher_name}')
            print(f'   课程组验收负责人: {course_leader}')
            
            # 检查表末尾端结构
            total_cell = ws.cell(row=19, column=1).value
            leader_cell = ws.cell(row=20, column=9).value
            date_cell = ws.cell(row=21, column=9).value
            note_cell = ws.cell(row=22, column=1).value
            
            has_total = total_cell == '合计'
            has_leader = leader_cell and '课程组负责人签字' in str(leader_cell)
            has_date = date_cell and '日期' in str(date_cell)
            has_note = note_cell and '温馨提醒' in str(note_cell)
            
            print(f'   合计行: {"✅" if has_total else "❌"}')
            print(f'   签字行: {"✅" if has_leader else "❌"}')
            print(f'   日期行: {"✅" if has_date else "❌"}')
            print(f'   温馨提醒: {"✅" if has_note else "❌"}')
            
            # 检查是否还有占位符
            has_placeholder = False
            for row_idx in range(1, ws.max_row + 1):
                for col_idx in range(1, ws.max_column + 1):
                    cell_value = ws.cell(row=row_idx, column=col_idx).value
                    if cell_value and ('{{' in str(cell_value) or '}}' in str(cell_value)):
                        has_placeholder = True
                        break
                if has_placeholder:
                    break
            
            print(f'   占位符替换: {"✅ 全部替换" if not has_placeholder else "❌ 还有占位符"}')
            
        except Exception as e:
            print(f'   ❌ 读取文件失败: {e}')
    else:
        print('   ❌ 未找到附件3文件')

print('\n' + '=' * 80)
print('✅ 验证完成！')
print('=' * 80)