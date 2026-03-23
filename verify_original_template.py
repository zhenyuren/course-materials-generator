from openpyxl import load_workbook

# 查看生成的附件3文件（使用原始模板）
wb = load_workbook('智能金融学院_未知教师/FIN3004A_大数据分析基础/大数据分析基础-课程组期初教学资料检查情况记录表-任渝.xlsx')
ws = wb.active

print('使用原始模板生成的附件3文件验证:')
print('=' * 80)

# 打印标题行
title = ws.cell(row=1, column=1).value
print(f'标题行: {title}')

# 打印表头和数据行
print('\n文件内容预览:')
for row_idx in range(1, 25):
    row_values = []
    for col_idx in range(1, 15):
        cell_value = ws.cell(row=row_idx, column=col_idx).value
        row_values.append(str(cell_value) if cell_value is not None else '')
    print(f'第{row_idx}行: ' + ' | '.join(row_values))

# 检查是否还有占位符
print('\n检查是否还有占位符:')
has_placeholder = False
for row_idx in range(1, ws.max_row + 1):
    for col_idx in range(1, ws.max_column + 1):
        cell_value = ws.cell(row=row_idx, column=col_idx).value
        if cell_value and ('{{' in str(cell_value) or '}}' in str(cell_value)):
            print(f'第{row_idx}行第{col_idx}列: {cell_value}')
            has_placeholder = True

if not has_placeholder:
    print('✅ 所有占位符已成功替换！')
else:
    print('❌ 还有占位符未替换！')

# 检查数据填充
print('\n数据填充检查:')
data_cells = [
    ('课程代码', 2, 5),
    ('课程名', 3, 5),
    ('开课单位', 4, 5),
    ('使用年级/层次/专业', 5, 5),
    ('教师（执笔人）', 7, 5),
    ('课程组验收负责人', 8, 5)
]
for label, col_idx, row_idx in data_cells:
    cell_value = ws.cell(row=row_idx, column=col_idx).value
    print(f'  {label}: {cell_value}')

# 检查表末尾端结构
print('\n表末尾端结构检查:')

# 合计行
total_cell = ws.cell(row=19, column=1)
print(f'第19行合计: {"✅ 存在" if total_cell.value == "合计" else "❌ 缺失"}')

# 总体评价意见
evaluation_cell = ws.cell(row=19, column=9)
print(f'总体评价意见: {"✅ 存在" if "总体评价意见" in str(evaluation_cell.value) else "❌ 缺失"}')

print('\n' + '=' * 80)
print('验证完成！')