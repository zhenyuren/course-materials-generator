from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# 创建工作簿
wb = Workbook()
ws = wb.active
ws.title = "Sheet1"

# 设置列宽（根据截图调整）
column_widths = {
    'A': 10,    # 序号
    'B': 12,    # 课程代码
    'C': 15,    # 课程名
    'D': 12,    # 开课单位
    'E': 20,    # 使用年级/层次/专业
    'F': 10,    # 归属校区
    'G': 12,    # 教师（执笔人）
    'H': 12,    # 课程组验收负责人
    'I': 12,    # 教学大纲符合度
    'J': 10,    # 教案撰写规范性
    'K': 10,    # 教学方式
    'L': 12,    # 教学进度及安排
    'M': 12,    # 课程设计和框架
    'N': 10     # 总分
}

for col_letter, width in column_widths.items():
    ws.column_dimensions[col_letter].width = width

# 设置行高
ws.row_dimensions[1].height = 30  # 标题行
ws.row_dimensions[2].height = 20  # 表头行
ws.row_dimensions[3].height = 18  # 评价指标行
ws.row_dimensions[4].height = 18  # 分数行
ws.row_dimensions[5].height = 18  # 数据行

# 创建边框样式
thin_border = Border(
    left=Side(border_style='thin', color='000000'),
    right=Side(border_style='thin', color='000000'),
    top=Side(border_style='thin', color='000000'),
    bottom=Side(border_style='thin', color='000000')
)

# 创建浅灰色背景
light_gray_fill = PatternFill(start_color='00DDDDDD', end_color='00DDDDDD', fill_type='solid')

# 设置标题行（合并A1:N1）
ws.merge_cells('A1:N1')
title_cell = ws['A1']
title_cell.value = '{{formattedSemester}}（{{courseName}}）课程组期初教学资料检查情况记录表'
title_cell.font = Font(name='宋体', size=16, bold=True)
title_cell.alignment = Alignment(horizontal='center', vertical='center')
title_cell.border = thin_border

# 合并第2、3、4行的各列（B-H列）
columns_to_merge = ['B', 'C', 'D', 'E', 'F', 'G', 'H']
for col_letter in columns_to_merge:
    ws.merge_cells(f'{col_letter}2:{col_letter}4')

# 设置表头行（合并后的单元格）
headers = ['课程代码', '课程名', '开课单位', '使用年级/层次/专业', '归属校区', '教师（执笔人）', '课程组验收负责人']
for i, header in enumerate(headers, 2):  # 从B列开始（索引2）
    cell = ws.cell(row=2, column=i, value=header)
    cell.font = Font(name='宋体', size=12, bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border

# 设置评价指标行（I-N列）
metrics = ['教学大纲符合度', '教案撰写规范性', '教学方式', '教学进度及安排', '课程设计和框架', '总分']
for i, metric in enumerate(metrics, 9):  # 从I列开始（索引9）
    cell = ws.cell(row=3, column=i, value=metric)
    cell.font = Font(name='宋体', size=11)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border
    cell.fill = light_gray_fill

# 设置分数行（I-N列）
scores = ['15分', '10分', '10分', '15分', '50分', '100分']
for i, score in enumerate(scores, 9):  # 从I列开始（索引9）
    cell = ws.cell(row=4, column=i, value=score)
    cell.font = Font(name='宋体', size=11)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border
    cell.fill = light_gray_fill

# 合并第2、3、4行A列
ws.merge_cells('A2:A4')
a2_cell = ws['A2']
a2_cell.value = '序号'
a2_cell.font = Font(name='宋体', size=12, bold=True)
a2_cell.alignment = Alignment(horizontal='center', vertical='center')
a2_cell.border = thin_border

# 合并第2、3、4行I-N列的表头
ws.merge_cells('I2:N2')
i2_cell = ws['I2']
i2_cell.value = '教案评价指标'
i2_cell.font = Font(name='宋体', size=12, bold=True)
i2_cell.alignment = Alignment(horizontal='center', vertical='center')
i2_cell.border = thin_border

# 设置数据行（占位符）
data = ['1', '{{courseCode}}', '{{courseName}}', '{{department}}', '{{applicableScope}}', '绵阳', '{{teacherName}}', '{{courseLeader}}', '12', '8', '8', '14', '43', '85']
for i, value in enumerate(data, 1):
    cell = ws.cell(row=5, column=i, value=value)
    cell.font = Font(name='宋体', size=11)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border

# 设置后续空行的边框
for row_idx in range(6, 20):
    for col_idx in range(1, 15):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.border = thin_border

# 添加合计行（第19行）
ws.merge_cells('A19:D19')
total_cell = ws['A19']
total_cell.value = '合计'
total_cell.font = Font(name='宋体', size=11, bold=True)
total_cell.alignment = Alignment(horizontal='center', vertical='center')
total_cell.border = thin_border

ws.merge_cells('E19:F19')
materials_cell = ws['E19']
materials_cell.value = '资料 份'
materials_cell.font = Font(name='宋体', size=11)
materials_cell.alignment = Alignment(horizontal='center', vertical='center')
materials_cell.border = thin_border

ws.merge_cells('G19:N19')
evaluation_cell = ws['G19']
evaluation_cell.value = '总体评价意见'
evaluation_cell.font = Font(name='宋体', size=11)
evaluation_cell.alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)
evaluation_cell.border = thin_border

# 添加课程组负责人签字行（第20行）
ws.merge_cells('A20:F20')
leader_cell = ws['A20']
leader_cell.value = '课程组负责人签字：'
leader_cell.font = Font(name='宋体', size=11)
leader_cell.alignment = Alignment(horizontal='left', vertical='center')
leader_cell.border = thin_border

# 添加日期行（第21行）
ws.merge_cells('A21:F21')
date_cell = ws['A21']
date_cell.value = '日期：'
date_cell.font = Font(name='宋体', size=11)
date_cell.alignment = Alignment(horizontal='left', vertical='center')
date_cell.border = thin_border

# 添加温馨提醒（第23-25行）
ws.merge_cells('A23:N23')
note1_cell = ws['A23']
note1_cell.value = '温馨提醒：1.此表一式两份，开课单位保存一份，教务处保存一份'
note1_cell.font = Font(name='宋体', size=11)
note1_cell.alignment = Alignment(horizontal='left', vertical='center')

ws.merge_cells('A24:N24')
note2_cell = ws['A24']
note2_cell.value = '2.60-70分为基本合格；70-80分为合格；80-90分为良好；90分以上为优秀'
note2_cell.font = Font(name='宋体', size=11)
note2_cell.alignment = Alignment(horizontal='left', vertical='center')

ws.merge_cells('A25:N25')
note3_cell = ws['A25']
note3_cell.value = '3.评价栏应对本课程组教学资料编写情况进行描述（50字以上）'
note3_cell.font = Font(name='宋体', size=11)
note3_cell.alignment = Alignment(horizontal='left', vertical='center')

# 保存模板文件
wb.save('期初资料/附件3 课程组教学资料检查情况记录表-完整版模板.xlsx')
print('✅ 完整格式的附件3模板已创建: 期初资料/附件3 课程组教学资料检查情况记录表-完整版模板.xlsx')