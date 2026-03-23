from openpyxl import load_workbook

# 查看生成的附件3文件
wb = load_workbook('智能金融学院_未知教师/FIN3004A_大数据分析基础/大数据分析基础-课程组期初教学资料检查情况记录表-任渝.xlsx')
ws = wb.active

print('附件3文件实际内容预览:')
print('=' * 80)

# 打印标题行
title = ws.cell(row=1, column=1).value
print(f'标题: {title}')

# 打印表头（第2行）
print('\n表头（第2行）:')
headers = []
for col_idx in range(1, 15):
    cell_value = ws.cell(row=2, column=col_idx).value
    headers.append(str(cell_value) if cell_value is not None else '')
print(' | '.join(headers))

# 打印评价指标行（第3行）
print('\n评价指标（第3行）:')
metrics = []
for col_idx in range(1, 15):
    cell_value = ws.cell(row=3, column=col_idx).value
    metrics.append(str(cell_value) if cell_value is not None else '')
print(' | '.join(metrics))

# 打印分数行（第4行）
print('\n分数（第4行）:')
scores = []
for col_idx in range(1, 15):
    cell_value = ws.cell(row=4, column=col_idx).value
    scores.append(str(cell_value) if cell_value is not None else '')
print(' | '.join(scores))

# 打印数据行（第5行）
print('\n数据行（第5行）:')
data = []
for col_idx in range(1, 15):
    cell_value = ws.cell(row=5, column=col_idx).value
    data.append(str(cell_value) if cell_value is not None else '')
print(' | '.join(data))

# 打印表末尾端内容
print('\n表末尾端内容:')

# 合计行（第19行）
total_line = []
for col_idx in range(1, 15):
    cell_value = ws.cell(row=19, column=col_idx).value
    total_line.append(str(cell_value) if cell_value is not None else '')
print('第19行: ' + ' | '.join(total_line))

# 课程组负责人签字（第20行）
leader_line = []
for col_idx in range(1, 15):
    cell_value = ws.cell(row=20, column=col_idx).value
    leader_line.append(str(cell_value) if cell_value is not None else '')
print('第20行: ' + ' | '.join(leader_line))

# 日期行（第21行）
date_line = []
for col_idx in range(1, 15):
    cell_value = ws.cell(row=21, column=col_idx).value
    date_line.append(str(cell_value) if cell_value is not None else '')
print('第21行: ' + ' | '.join(date_line))

# 温馨提醒（第23-25行）
print('\n温馨提醒:')
for row_idx in range(23, 26):
    note_text = ws.cell(row=row_idx, column=1).value
    print(f'第{row_idx}行: {note_text}')

# 检查合并单元格的实际效果
print('\n' + '=' * 80)
print('合并单元格实际效果检查:')

# 检查B2:B4合并
b2_cell = ws.cell(row=2, column=2)
b3_cell = ws.cell(row=3, column=2)
b4_cell = ws.cell(row=4, column=2)
print(f'B2值: {b2_cell.value}, B3值: {b3_cell.value}, B4值: {b4_cell.value}')
print(f'B2是否为合并单元格的左上角: {b2_cell.value is not None and b3_cell.value is None and b4_cell.value is None}')

# 检查I2:N2合并
i2_cell = ws.cell(row=2, column=9)
j2_cell = ws.cell(row=2, column=10)
k2_cell = ws.cell(row=2, column=11)
print(f'I2值: {i2_cell.value}, J2值: {j2_cell.value}, K2值: {k2_cell.value}')
print(f'I2是否为合并单元格的左上角: {i2_cell.value is not None and j2_cell.value is None and k2_cell.value is None}')

print('\n' + '=' * 80)
print('预览完成！')