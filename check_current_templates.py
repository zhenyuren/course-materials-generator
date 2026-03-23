from openpyxl import load_workbook

# 查看当前附件1模板
wb1 = load_workbook('期初资料/附件1 教学大纲的审核汇总表-大数据信息分析_任渝.xlsx')
ws1 = wb1.active

print('当前附件1模板内容:')
print('=' * 80)
for row_idx in range(1, min(ws1.max_row + 1, 8)):
    row_values = []
    for col_idx in range(1, min(ws1.max_column + 1, 15)):
        cell_value = ws1.cell(row=row_idx, column=col_idx).value
        row_values.append(str(cell_value) if cell_value is not None else 'None')
    print(f"行 {row_idx}: {', '.join(row_values)}")

print('\n' + '=' * 80)
print('合并单元格信息:')
merged_list = list(ws1.merged_cells.ranges)
for i, merged_cell in enumerate(merged_list[:10], 1):
    print(f"  {i}. {merged_cell}")
if len(merged_list) > 10:
    print(f"  ... 还有 {len(merged_list) - 10} 个合并单元格")