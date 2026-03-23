from openpyxl import load_workbook

# 读取生成的文件
wb = load_workbook('智能金融学院_未知教师/FIN3004A_大数据分析基础/大数据分析基础-课程组期初教学资料检查情况记录表-任渝.xlsx')
ws = wb.active

print('生成文件的实际内容:')
print('=' * 80)

# 打印所有内容
for row_idx in range(1, 25):
    row_values = []
    for col_idx in range(1, 15):
        cell_value = ws.cell(row=row_idx, column=col_idx).value
        row_values.append(str(cell_value) if cell_value is not None else '')
    print(f'第{row_idx}行: ' + ' | '.join(row_values))

# 检查合并单元格
print('\n合并单元格:')
merged_list = list(ws.merged_cells.ranges)
for merged_cell in merged_list:
    print(f'  {merged_cell}')

# 检查单元格格式
print('\n关键单元格格式:')
# 标题
title_cell = ws.cell(row=1, column=1)
print(f'标题字体: {title_cell.font.name}, 大小: {title_cell.font.size}, 加粗: {title_cell.font.bold}')

# 表头
for col_idx in range(1, 9):
    cell = ws.cell(row=2, column=col_idx)
    print(f'表头列{col_idx}: 字体={cell.font.name}, 大小={cell.font.size}, 加粗={cell.font.bold}')

# 评价指标
for col_idx in range(9, 15):
    cell = ws.cell(row=3, column=col_idx)
    print(f'评价指标列{col_idx}: 字体={cell.font.name}, 大小={cell.font.size}, 背景色={cell.fill.fgColor.index}')

# 分数
for col_idx in range(9, 15):
    cell = ws.cell(row=4, column=col_idx)
    print(f'分数列{col_idx}: 字体={cell.font.name}, 大小={cell.font.size}, 背景色={cell.fill.fgColor.index}')