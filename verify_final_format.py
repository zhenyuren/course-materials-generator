from openpyxl import load_workbook
import json

# 读取JSON数据
with open('json/任渝_课程信息_2026年03月21日.json', 'r', encoding='utf-8') as f:
    data = json.load(f)
    courses = data.get('courses', [])

print('最终格式验证:')
print('=' * 80)

# 验证每门课程的生成文件
for course in courses:
    course_code = course.get('courseCode', '')
    course_name = course.get('courseName', '')
    teacher_name = course.get('teacherName', '')
    current_date = course.get('currentDate', '')
    
    # 跳过重复课程
    if course_code == 'FIN4011B' and course_name == '跨境电商大数据统计与分析':
        continue
    if course_code == 'FIN3005A' and course_name == '大数据处理技术':
        continue
    
    print(f'\n课程: {course_name} ({course_code})')
    
    # 读取生成的文件
    file_path = f'智能金融学院_未知教师/{course_code}_{course_name}/{course_name}-课程组期初教学资料检查情况记录表-任渝.xlsx'
    try:
        wb = load_workbook(file_path)
        ws = wb.active
        
        # 获取数据行信息
        course_code_cell = ws.cell(row=5, column=2).value
        course_name_cell = ws.cell(row=5, column=3).value
        department_cell = ws.cell(row=5, column=4).value
        grade_major_cell = ws.cell(row=5, column=5).value
        campus_cell = ws.cell(row=5, column=6).value
        teacher_cell = ws.cell(row=5, column=7).value
        
        # 获取分数
        outline_score = ws.cell(row=5, column=9).value
        norm_score = ws.cell(row=5, column=10).value
        objective_score = ws.cell(row=5, column=11).value
        method_score = ws.cell(row=5, column=12).value
        process_score = ws.cell(row=5, column=13).value
        total_score = ws.cell(row=5, column=14).value
        
        # 获取资料份数
        materials_cell = ws.cell(row=19, column=5).value
        
        # 获取课程组负责人签字
        leader_cell = ws.cell(row=20, column=7).value
        
        # 获取日期
        date_cell = ws.cell(row=21, column=9).value
        
        # 获取总体评价意见
        evaluation_cell = ws.cell(row=19, column=8).value
        
        print('基本信息:')
        print(f'课程代码: {course_code_cell}')
        print(f'课程名: {course_name_cell}')
        print(f'开课单位: {department_cell}')
        print(f'使用年级/层次/专业: {grade_major_cell}')
        print(f'归属校区: {campus_cell}')
        print(f'教师（执笔人）: {teacher_cell}')
        
        print('\n分数信息:')
        print(f'教学大纲符合度: {outline_score}')
        print(f'撰写规范性: {norm_score}')
        print(f'教学目标设计: {objective_score}')
        print(f'教学方法设计: {method_score}')
        print(f'教学过程设计: {process_score}')
        print(f'总分: {total_score}')
        
        print('\n其他信息:')
        print(f'资料份数: {materials_cell}')
        print(f'课程组负责人签字: {leader_cell}')
        print(f'日期: {date_cell}')
        print(f'总体评价意见: {evaluation_cell}')
        
        # 验证分数
        valid = True
        
        if outline_score >= 15:
            print('❌ 教学大纲符合度分数必须小于15')
            valid = False
        
        if norm_score >= 10:
            print('❌ 撰写规范性分数必须小于10')
            valid = False
        
        if objective_score >= 15:
            print('❌ 教学目标设计分数必须小于15')
            valid = False
        
        if method_score >= 10:
            print('❌ 教学方法设计分数必须小于10')
            valid = False
        
        if process_score >= 50:
            print('❌ 教学过程设计分数必须小于50')
            valid = False
        
        if total_score < 86 or total_score > 92:
            print('❌ 总分必须在86-92之间')
            valid = False
        
        # 验证资料份数
        if '资料' not in materials_cell or '份' not in materials_cell:
            print('❌ 资料份数格式不正确')
            valid = False
        
        # 验证课程组负责人签字
        if '课程组负责人签字' not in leader_cell:
            print('❌ 课程组负责人签字格式不正确')
            valid = False
        
        # 验证日期（日期在I列，显示的是纯日期值）
        if date_cell != current_date:
            print('❌ 日期格式不正确')
            valid = False
        
        # 验证总体评价意见
        expected_evaluation = '满足教学大纲要求，课堂时间分配合理，教学实施设计上有的放矢，教学重点、难点把握得当，课程组审核合格。'
        if evaluation_cell != expected_evaluation:
            print('❌ 总体评价意见不正确')
            valid = False
        
        if valid:
            print('\n✅ 所有验证通过！')
            
    except FileNotFoundError:
        print(f'❌ 文件不存在: {file_path}')
    except Exception as e:
        print(f'❌ 验证出错: {e}')

print('\n' + '=' * 80)
print('验证完成！')