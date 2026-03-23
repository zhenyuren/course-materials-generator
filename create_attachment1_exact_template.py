from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# 创建工作簿
wb = Workbook()
ws = wb.active

# 创建边框样式
thin_border = Border(
    left=Side(border_style='thin', color='000000'),
    right=Side(border_style='thin', color='000000'),
    top=Side(border_style='thin', color='000000'),
    bottom=Side(border_style='thin', color='000000')
)

# 创建无边框样式
no_border = Border(
    left=None,
    right=None,
    top=None,
    bottom=None
)

# 创建浅灰色背景
light_gray_fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')

# 设置列宽（根据附件1.xls的实际宽度调整）
column_widths = [5, 12, 20, 12, 10, 22, 15, 15, 15, 15, 12, 15, 20, 10]
for i, width in enumerate(column_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = width

# 设置标题行（合并A1:N1）
ws.merge_cells('A1:N1')
title_cell = ws['A1']
title_cell.value = '{{formattedSemester}}学期各开课单位教学大纲的审核汇总表'
title_cell.font = Font(name='宋体', size=16, bold=True)
title_cell.alignment = Alignment(horizontal='center', vertical='center')
title_cell.border = thin_border

# 设置表头（第2行）- 先不合并
headers_row2 = [
    '序号',
    '课程代码',
    '课程名',
    '开课单位',
    '所属校区',
    '使用年级/层次/专业',
    '执笔人',
    '验收负责人',
    '教学大纲评价指标',
    '',
    '',
    '',
    '',
    ''
]
for i, header in enumerate(headers_row2, 1):
    cell = ws.cell(row=2, column=i, value=header)
    cell.font = Font(name='宋体', size=12, bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border
    # 前8列有灰色背景
    if i <= 8:
        cell.fill = light_gray_fill

# 设置评价指标行（第3行）
metrics_row3 = [
    '',
    '',
    '',
    '',
    '',
    '',
    '',
    '',
    '',
    '',
    '人才培养目标符合度',
    '大纲撰写规范性',
    '教学方式',
    '教学进度及安排',
    '课程设计和框架',
    '总分'
]
for i, metric in enumerate(metrics_row3, 1):
    cell = ws.cell(row=3, column=i, value=metric)
    cell.font = Font(name='宋体', size=10)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border
    # 前8列有灰色背景
    if i <= 8:
        cell.fill = light_gray_fill

# 设置分数行（第4行）
scores_row4 = [
    '',
    '',
    '',
    '',
    '',
    '',
    '',
    '',
    '',
    '',
    '15分',
    '10分',
    '10分',
    '15分',
    '50分',
    '100分'
]
for i, score in enumerate(scores_row4, 1):
    cell = ws.cell(row=4, column=i, value=score)
    cell.font = Font(name='宋体', size=10)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border
    # 前8列有灰色背景
    if i <= 8:
        cell.fill = light_gray_fill

# 设置数据行占位符（第5行）
data_row5 = ['1', '{{courseCode}}', '{{courseName}}', '{{department}}', '{{campus}}', '{{gradeMajor}}', '{{teacherName}}', '{{courseLeader}}', '', '', '', '', '', '']
for i, value in enumerate(data_row5, 1):
    cell = ws.cell(row=5, column=i, value=value)
    cell.font = Font(name='宋体', size=11)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border

# 设置空行（第6-17行）- 序号2-13
for row_idx in range(6, 18):
    # 设置序号
    cell = ws.cell(row=row_idx, column=1, value=row_idx - 4)
    cell.font = Font(name='宋体', size=11)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border
    # 设置其他列的边框
    for col_idx in range(2, 15):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.font = Font(name='宋体', size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

# 设置行高
ws.row_dimensions[18].height = 40
ws.row_dimensions[19].height = 25
ws.row_dimensions[20].height = 25

# 添加合计行（第18行）
total_cell = ws.cell(row=18, column=1, value='合计')
total_cell.font = Font(name='宋体', size=11, bold=True)
total_cell.alignment = Alignment(horizontal='center', vertical='center')
total_cell.border = thin_border

materials_cell = ws.cell(row=18, column=2, value='资料 份')
materials_cell.font = Font(name='宋体', size=11)
materials_cell.alignment = Alignment(horizontal='center', vertical='center')
materials_cell.border = thin_border

evaluation_cell = ws.cell(row=18, column=9, value='总体评价意见')
evaluation_cell.font = Font(name='宋体', size=11)
evaluation_cell.alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)
evaluation_cell.border = thin_border

evaluation_text_cell = ws.cell(row=18, column=10, value='教学大纲符合课程设置要求，教学进度与教学内容与教学大纲设置适配度较高。')
evaluation_text_cell.font = Font(name='宋体', size=11)
evaluation_text_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
evaluation_text_cell.border = thin_border

# 添加开课单位负责人签字行（第19行）
leader_cell = ws.cell(row=19, column=9, value='开课单位负责人签字：')
leader_cell.font = Font(name='宋体', size=11)
leader_cell.alignment = Alignment(horizontal='left', vertical='center')
leader_cell.border = no_border

# 添加日期行（第20行）
date_cell = ws.cell(row=20, column=9, value='日期：')
date_cell.font = Font(name='宋体', size=11)
date_cell.alignment = Alignment(horizontal='left', vertical='center')
date_cell.border = no_border

# 现在进行合并单元格操作
# 合并第2行I-N列（教学大纲评价指标）
ws.merge_cells('I2:N2')

# 合并合计行
ws.merge_cells('A18:H18')
ws.merge_cells('I18:N18')

# 合并开课单位负责人签字后面的格子（J-N列）
ws.merge_cells('J19:N19')

# 合并日期后面的格子（J-N列）
ws.merge_cells('J20:N20')

# 保存文件
wb.save('期初资料/附件1 教学大纲审核汇总表-完全一致模板.xlsx')
print('✅ 附件1完全一致模板已创建: 期初资料/附件1 教学大纲审核汇总表-完全一致模板.xlsx')
