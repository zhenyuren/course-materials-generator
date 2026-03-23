from openpyxl import load_workbook

# 查看附件3模板结构
wb = load_workbook('期初资料/附件3 课程组教学资料检查情况记录表.xlsx')
ws = wb.active
print('附件3模板结构:')
print(f'总行数: {ws.max_row}')
print(f'总列数: {ws.max_column}')
print('前5行内容:')
for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
    print(row)

print('\n' + '='*50 + '\n')

# 查看附件1模板结构
wb = load_workbook('期初资料/附件1 教学大纲的审核汇总表-大数据信息分析_任渝.xlsx')
ws = wb.active
print('附件1模板结构:')
print(f'总行数: {ws.max_row}')
print(f'总列数: {ws.max_column}')
print('前5行内容:')
for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
    print(row)