import xlrd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

def convert_xls_to_xlsx_preserve_format(xls_path, xlsx_path):
    """读取xls文件并创建xlsx文件，尽可能保留格式"""
    
    # 读取xls文件
    workbook_xls = xlrd.open_workbook(xls_path, formatting_info=True)
    sheet_xls = workbook_xls.sheet_by_index(0)
    
    # 创建新的xlsx工作簿
    workbook_xlsx = Workbook()
    sheet_xlsx = workbook_xlsx.active
    
    # 设置列宽
    for col_idx in range(sheet_xls.ncols):
        width = sheet_xls.colinfo(col_idx).width
        if width > 0:
            sheet_xlsx.column_dimensions[get_column_letter(col_idx + 1)].width = width
    
    # 复制单元格内容
    for row_idx in range(sheet_xls.nrows):
        for col_idx in range(sheet_xls.ncols):
            cell_value = sheet_xls.cell_value(row_idx, col_idx)
            cell_type = sheet_xls.cell_type(row_idx, col_idx)
            
            # 处理不同类型的值
            if cell_type == xlrd.XL_CELL_NUMBER:
                # 如果是整数，保存为整数
                if cell_value == int(cell_value):
                    cell_value = int(cell_value)
            
            # 设置单元格值
            sheet_xlsx.cell(row=row_idx + 1, column=col_idx + 1, value=cell_value)
    
    # 保存文件
    workbook_xlsx.save(xlsx_path)
    print(f"✅ 成功转换并保留格式: {xls_path} -> {xlsx_path}")

# 转换原始xls文件
xls_file = '期初资料/附件3 课程组教学资料检查情况记录表.xls'
xlsx_file = '期初资料/附件3 课程组教学资料检查情况记录表-原始格式.xlsx'

convert_xls_to_xlsx_preserve_format(xls_file, xlsx_file)