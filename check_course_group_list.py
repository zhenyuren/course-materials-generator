from openpyxl import load_workbook

wb = load_workbook('期初资料1/2026年春季学期课程组名单-智能金融学院.xlsx')
ws = wb.active

print('列名:')
for cell in ws[1]:
    print(f"列 {cell.column_letter}: {cell.value}")

print('\n前5行数据:')
for row in ws.iter_rows(min_row=2, max_row=6, values_only=True):
    print(row)
