from openpyxl import load_workbook
import os

# 测试附件3模板
template_path = "期初资料/附件3 课程组教学资料检查情况记录表.xlsx"
output_path = "test_attachment3.xlsx"

# 复制模板文件
import shutil
shutil.copy(template_path, output_path)

# 加载模板并填充数据
wb = load_workbook(output_path)
ws = wb.active

# 填充数据到第5行
ws.cell(row=5, column=2, value="TEST001")  # 课程代码
ws.cell(row=5, column=3, value="测试课程")  # 课程名称
ws.cell(row=5, column=4, value="智能金融学院")  # 开课单位
ws.cell(row=5, column=6, value="东区24级本数字经济01班")  # 使用年级/层次/专业
ws.cell(row=5, column=7, value="任渝")  # 教师（执笔人）
ws.cell(row=5, column=8, value="任渝")  # 课程组验收负责人

# 保存文件
wb.save(output_path)

print(f"测试文件已生成: {output_path}")

# 分析生成的文件
wb = load_workbook(output_path)
ws = wb.active

print("\n生成文件的内容:")
for row in ws.iter_rows(min_row=1, max_row=6, values_only=True):
    print(row)

# 清理测试文件
os.remove(output_path)