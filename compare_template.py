from openpyxl import load_workbook

# 读取原始模板
wb_template = load_workbook('期初资料/附件3 课程组教学资料检查情况记录表_原始.xlsx')
ws_template = wb_template.active

# 读取生成的文件
wb_generated = load_workbook('智能金融学院_未知教师/FIN3004A_大数据分析基础/大数据分析基础-课程组期初教学资料检查情况记录表-任渝.xlsx')
ws_generated = wb_generated.active

print('模板文件 vs 生成文件对比:')
print('=' * 80)

print('\n第2行表头对比:')
print('模板文件:')
for col_idx in range(1, 15):
    value = ws_template.cell(row=2, column=col_idx).value
    print(f'  列{col_idx}: {value}')

print('\n生成文件:')
for col_idx in range(1, 15):
    value = ws_generated.cell(row=2, column=col_idx).value
    print(f'  列{col_idx}: {value}')

print('\n第3行评价指标对比:')
print('模板文件:')
for col_idx in range(1, 15):
    value = ws_template.cell(row=3, column=col_idx).value
    print(f'  列{col_idx}: {value}')

print('\n生成文件:')
for col_idx in range(1, 15):
    value = ws_generated.cell(row=3, column=col_idx).value
    print(f'  列{col_idx}: {value}')

print('\n第4行分数对比:')
print('模板文件:')
for col_idx in range(1, 15):
    value = ws_template.cell(row=4, column=col_idx).value
    print(f'  列{col_idx}: {value}')

print('\n生成文件:')
for col_idx in range(1, 15):
    value = ws_generated.cell(row=4, column=col_idx).value
    print(f'  列{col_idx}: {value}')

# 检查表末尾端结构
print('\n表末尾端结构对比:')
print('模板文件第19行:')
for col_idx in range(1, 15):
    value = ws_template.cell(row=19, column=col_idx).value
    print(f'  列{col_idx}: {value}')

print('\n生成文件第19行:')
for col_idx in range(1, 15):
    value = ws_generated.cell(row=19, column=col_idx).value
    print(f'  列{col_idx}: {value}')