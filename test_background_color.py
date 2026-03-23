from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# 创建测试工作簿
wb = Workbook()
ws = wb.active

# 创建边框样式
thin_border = Border(
    left=Side(border_style='thin', color='000000'),
    right=Side(border_style='thin', color='000000'),
    top=Side(border_style='thin', color='000000'),
    bottom=Side(border_style='thin', color='000000')
)

# 创建浅灰色背景
light_gray_fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')

# 设置一个测试单元格
cell = ws.cell(row=1, column=1, value='测试')
cell.font = Font(name='宋体', size=11)
cell.alignment = Alignment(horizontal='center', vertical='center')
cell.border = thin_border
cell.fill = light_gray_fill

# 保存测试文件
wb.save('test_background.xlsx')
print('测试文件已创建: test_background.xlsx')

# 重新读取并检查背景色
wb2 = Workbook()
ws2 = wb2.active
ws2['A1'] = '测试2'
ws2['A1'].fill = light_gray_fill
wb2.save('test_background2.xlsx')
print('测试文件2已创建: test_background2.xlsx')

# 直接创建模板文件来测试
wb3 = Workbook()
ws3 = wb3.active

# 设置评价指标行（带背景色）
for i in range(9, 15):
    cell = ws3.cell(row=3, column=i, value=f'指标{i-8}')
    cell.font = Font(name='宋体', size=11)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border
    cell.fill = light_gray_fill

# 设置分数行（带背景色）
for i in range(9, 15):
    cell = ws3.cell(row=4, column=i, value=f'{i*10}分')
    cell.font = Font(name='宋体', size=11)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border
    cell.fill = light_gray_fill

wb3.save('test_background3.xlsx')
print('测试文件3已创建: test_background3.xlsx')

# 检查颜色代码
print(f'背景色代码: {light_gray_fill.fgColor.index}')