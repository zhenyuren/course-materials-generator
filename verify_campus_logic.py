from openpyxl import load_workbook

# 验证不同类型的课程数据填充
print("验证不同类型课程的校区填充逻辑：")
print("-" * 50)

# 1. 有具体区域信息的课程（东区）
file_path1 = "智能金融学院_未知教师/FIN0008B_财经公文写作/财经公文写作-教学大纲审核汇总表-未知教师.xlsx"
wb1 = load_workbook(file_path1)
ws1 = wb1.active

print("财经公文写作（有东区信息）：")
print(f"所属校区: {ws1.cell(row=5, column=5).value}")
print(f"使用年级/层次/专业: {ws1.cell(row=5, column=6).value}")
print()

# 2. 没有具体区域信息的课程
file_path2 = "智能金融学院_未知教师/FIN3005A_大数据处理技术/大数据处理技术-教学大纲审核汇总表-未知教师.xlsx"
wb2 = load_workbook(file_path2)
ws2 = wb2.active

print("大数据处理技术（无区域信息）：")
print(f"所属校区: {ws2.cell(row=5, column=5).value}")
print(f"使用年级/层次/专业: {ws2.cell(row=5, column=6).value}")
print()

# 3. 验证附件3的校区逻辑
file_path3 = "智能金融学院_未知教师/FIN0008B_财经公文写作/财经公文写作-课程组期初教学资料检查情况记录表-未知教师.xlsx"
wb3 = load_workbook(file_path3)
ws3 = wb3.active

print("附件3 - 财经公文写作（有东区信息）：")
print(f"归属校区: {ws3.cell(row=5, column=6).value}")
print(f"使用年级/层次/专业: {ws3.cell(row=5, column=5).value}")
print()

# 4. 验证附件3的默认西区逻辑
file_path4 = "智能金融学院_未知教师/FIN3005A_大数据处理技术/大数据处理技术-课程组期初教学资料检查情况记录表-未知教师.xlsx"
wb4 = load_workbook(file_path4)
ws4 = wb4.active

print("附件3 - 大数据处理技术（无区域信息）：")
print(f"归属校区: {ws4.cell(row=5, column=6).value}")
print(f"使用年级/层次/专业: {ws4.cell(row=5, column=5).value}")
