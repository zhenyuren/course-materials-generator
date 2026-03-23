import json
import os
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from docx import Document
from docx.shared import Inches

class SemesterMaterialsGenerator:
    def __init__(self):
        self.template_dir = '期初资料1'
        self.courses = []
        self.course_group_leaders = {}
        self.metadata = {}
    
    def load_course_group_data(self):
        """加载课程组名单数据"""
        file_path = os.path.join(self.template_dir, "2026年春季学期课程组名单-智能金融学院.xlsx")
        if not os.path.exists(file_path):
            print(f"❌ 课程组名单文件不存在: {file_path}")
            return False
        
        try:
            wb = load_workbook(file_path)
            ws = wb.active
            
            current_course = None
            for row in ws.iter_rows(min_row=2, values_only=True):
                campus, course_code, course_name, credits, leader, member, remark = row
                
                if course_code is not None:
                    current_course = {
                        'courseCode': course_code,
                        'courseName': course_name,
                        'campus': campus,
                        'credits': credits,
                        'courseLeader': leader,
                        'department': '智能金融学院',
                        'courseMembers': [leader] if leader else []
                    }
                    self.course_group_leaders[course_code] = leader
                    self.courses.append(current_course)
                elif current_course is not None and member is not None:
                    current_course['courseMembers'].append(member)
            
            print(f"✅ 成功加载 {len(self.courses)} 条课程组数据")
            return True
        except Exception as e:
            print(f"❌ 加载课程组数据失败: {e}")
            return False
    
    def load_course_info(self):
        """加载任渝的课程信息"""
        file_path = os.path.join(self.template_dir, "任渝_课程信息_2026年03月21日.json")
        if not os.path.exists(file_path):
            print(f"❌ 课程信息文件不存在: {file_path}")
            return False
        
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            
            if isinstance(data, dict) and 'courses' in data:
                self.courses.extend(data['courses'])
            else:
                self.courses.extend(data)
            
            print(f"✅ 成功加载任渝的课程信息")
            return True
        except Exception as e:
            print(f"❌ 加载课程信息失败: {e}")
            return False
    
    def get_course_leader(self, course_code):
        """根据课程代码获取课程负责人"""
        if course_code in self.course_group_leaders:
            return self.course_group_leaders[course_code]
        
        for course in self.courses:
            if course.get('courseCode') == course_code:
                return course.get('courseLeader', '')
        return ''
    
    def generate_attachment3(self, course):
        """生成附件3：课程组期初教学资料检查情况记录表"""
        template_file = os.path.join(self.template_dir, "附件3 课程组教学资料检查情况记录表-完全一致模板.xlsx")
        if not os.path.exists(template_file):
            print(f"❌ 附件3模板文件不存在: {template_file}")
            return None
        
        try:
            wb = load_workbook(template_file)
            ws = wb.active
            
            course_code = course.get('courseCode', '')
            course_name = course.get('courseName', '')
            department = course.get('department', '')
            applicable_scope = course.get('applicableScope', '')
            teacher_name = course.get('teacherName', '')
            course_leader = self.get_course_leader(course_code)
            
            # 设置标题行（使用占位符转换方法）
            title_cell = ws.cell(row=1, column=1)
            title_cell.value = f"{course.get('formattedSemester', '2026春')}学期（{course_name}）课程组期初教学资料检查情况记录表"
            
            # 解析applicableScope提取归属校区和使用年级/层次/专业
            campus = ""
            grade_major = applicable_scope
            
            if applicable_scope.startswith(('东区', '绵阳校区', '德阳校区')):
                if applicable_scope.startswith('东区'):
                    campus = '东区'
                    grade_major = applicable_scope[2:]
                elif applicable_scope.startswith('绵阳校区'):
                    campus = '绵阳校区'
                    grade_major = applicable_scope[5:]
                elif applicable_scope.startswith('德阳校区'):
                    campus = '德阳校区'
                    grade_major = applicable_scope[5:]
            else:
                campus = '西区'
                grade_major = applicable_scope
            
            # 在第5行开始填充数据
            ws.cell(row=5, column=1, value='1')
            ws.cell(row=5, column=2, value=course_code)
            ws.cell(row=5, column=3, value=course_name)
            ws.cell(row=5, column=4, value=department)
            ws.cell(row=5, column=5, value=grade_major)
            ws.cell(row=5, column=6, value=campus)
            ws.cell(row=5, column=7, value=teacher_name)
            ws.cell(row=5, column=8, value=course_leader)
            
            # 生成分数，满足总分86-92之间
            import random
            while True:
                goal_score = random.randint(12, 14)
                norm_score = random.randint(8, 9)
                method_score = random.randint(8, 9)
                schedule_score = random.randint(12, 14)
                design_score = random.randint(42, 48)
                
                total = goal_score + norm_score + method_score + schedule_score + design_score
                
                if 86 <= total <= 92:
                    break
            
            ws.cell(row=5, column=9, value=goal_score)
            ws.cell(row=5, column=10, value=norm_score)
            ws.cell(row=5, column=11, value=method_score)
            ws.cell(row=5, column=12, value=schedule_score)
            ws.cell(row=5, column=13, value=design_score)
            ws.cell(row=5, column=14, value=total)
            
            # 更新资料份数
            teacher_count = 0
            for c in self.courses:
                if c.get('courseCode') == course_code:
                    teacher_count += 1
            
            materials_cell = ws.cell(row=19, column=5)
            materials_cell.value = f'资料 {teacher_count} 份'
            
            # 更新日期
            current_date = self.metadata.get('date', '')
            if current_date:
                date_cell = ws.cell(row=21, column=8)
                date_cell.value = f"日期：{current_date}"
            
            # 添加电子签名图片（课程组负责人签字后面）
            signature_image_path = os.path.join(self.template_dir, "任渝.png")
            
            if os.path.exists(signature_image_path):
                try:
                    img = Image(signature_image_path)
                    img.width = 100
                    img.height = 40
                    ws.add_image(img, 'M20')
                    print(f"✅ 电子签名已添加到 {course_name} 的附件3")
                except Exception as e:
                    print(f"⚠️ 添加电子签名失败: {e}")
            
            return wb
        except Exception as e:
            print(f"❌ 生成附件3失败: {e}")
            return None
    
    def generate_attachment1(self, course):
        """生成附件1：教学大纲审核汇总表"""
        template_file = os.path.join(self.template_dir, "附件1 教学大纲审核汇总表-完全一致模板.xlsx")
        if not os.path.exists(template_file):
            print(f"❌ 附件1模板文件不存在: {template_file}")
            return None
        
        try:
            wb = load_workbook(template_file)
            ws = wb.active
            
            course_code = course.get('courseCode', '')
            course_name = course.get('courseName', '')
            course_leader = self.get_course_leader(course_code)
            department = course.get('department', '')
            applicable_scope = course.get('applicableScope', '')
            teacher_name = course.get('teacherName', '')
            
            # 设置标题行（使用占位符转换方法）
            title_cell = ws.cell(row=1, column=1)
            title_cell.value = f"{course.get('formattedSemester', '2026春')}学期（{course_name}）教学大纲审核汇总表"
            
            # 解析applicableScope提取归属校区和使用年级/层次/专业
            campus = ""
            grade_major = applicable_scope
            
            if applicable_scope.startswith(('东区', '绵阳校区', '德阳校区')):
                if applicable_scope.startswith('东区'):
                    campus = '东区'
                    grade_major = applicable_scope[2:]
                elif applicable_scope.startswith('绵阳校区'):
                    campus = '绵阳校区'
                    grade_major = applicable_scope[5:]
                elif applicable_scope.startswith('德阳校区'):
                    campus = '德阳校区'
                    grade_major = applicable_scope[5:]
            else:
                campus = '西区'
                grade_major = applicable_scope
            
            # 在第5行开始填充数据
            ws.cell(row=5, column=1, value='1')
            ws.cell(row=5, column=2, value=course_code)
            ws.cell(row=5, column=3, value=course_name)
            ws.cell(row=5, column=4, value=department)
            ws.cell(row=5, column=5, value=campus)
            ws.cell(row=5, column=6, value=grade_major)
            ws.cell(row=5, column=7, value=teacher_name)
            ws.cell(row=5, column=8, value=course_leader)
            
            # 生成分数，满足总分86-92之间
            import random
            while True:
                goal_score = random.randint(12, 14)
                norm_score = random.randint(8, 9)
                method_score = random.randint(8, 9)
                schedule_score = random.randint(12, 14)
                design_score = random.randint(42, 48)
                
                total = goal_score + norm_score + method_score + schedule_score + design_score
                
                if 86 <= total <= 92:
                    break
            
            ws.cell(row=5, column=9, value=goal_score)
            ws.cell(row=5, column=10, value=norm_score)
            ws.cell(row=5, column=11, value=method_score)
            ws.cell(row=5, column=12, value=schedule_score)
            ws.cell(row=5, column=13, value=design_score)
            ws.cell(row=5, column=14, value=total)
            
            # 计算资料份数
            teacher_count = 0
            for c in self.courses:
                if c.get('courseCode') == course_code:
                    teacher_count += 1
            
            materials_cell = ws.cell(row=18, column=1)
            materials_cell.value = f'资料 {teacher_count} 份'
            
            # 更新日期
            current_date = self.metadata.get('date', '')
            if current_date:
                date_cell = ws.cell(row=20, column=10)
                date_cell.value = f"日期：{current_date}"
            
            # 添加电子签名图片
            signature_image_path = os.path.join(self.template_dir, "任渝.png")
            
            if os.path.exists(signature_image_path):
                try:
                    img = Image(signature_image_path)
                    img.width = 100
                    img.height = 40
                    ws.add_image(img, 'J19')
                    print(f"✅ 电子签名已添加到 {course_name} 的附件1")
                except Exception as e:
                    print(f"⚠️ 添加电子签名失败: {e}")
            
            return wb
        except Exception as e:
            print(f"❌ 生成附件1失败: {e}")
            return None
    
    def generate_syllabus(self, course):
        """生成教学大纲"""
        template_file = os.path.join(self.template_dir, "1 课程教学大纲基础模版.docx")
        if not os.path.exists(template_file):
            print(f"❌ 教学大纲模板文件不存在: {template_file}")
            return None
        
        try:
            doc = Document(template_file)
            
            course_code = course.get('courseCode', '')
            course_name = course.get('courseName', '')
            teacher_name = course.get('teacherName', '')
            department = course.get('department', '')
            credits = course.get('credits', '')
            total_hours = course.get('totalHours', '')
            course_nature = course.get('courseNature', '')
            applicable_scope = course.get('applicableScope', '')
            english_name = course.get('englishName', '')
            
            # 替换占位符
            for paragraph in doc.paragraphs:
                if '{{courseCode}}' in paragraph.text:
                    paragraph.text = paragraph.text.replace('{{courseCode}}', course_code)
                if '{{courseName}}' in paragraph.text:
                    paragraph.text = paragraph.text.replace('{{courseName}}', course_name)
                if '{{teacherName}}' in paragraph.text:
                    paragraph.text = paragraph.text.replace('{{teacherName}}', teacher_name)
                if '{{department}}' in paragraph.text:
                    paragraph.text = paragraph.text.replace('{{department}}', department)
                if '{{credits}}' in paragraph.text:
                    paragraph.text = paragraph.text.replace('{{credits}}', credits)
                if '{{totalHours}}' in paragraph.text:
                    paragraph.text = paragraph.text.replace('{{totalHours}}', total_hours)
                if '{{courseNature}}' in paragraph.text:
                    paragraph.text = paragraph.text.replace('{{courseNature}}', course_nature)
                if '{{applicableScope}}' in paragraph.text:
                    paragraph.text = paragraph.text.replace('{{applicableScope}}', applicable_scope)
                if '{{englishName}}' in paragraph.text:
                    paragraph.text = paragraph.text.replace('{{englishName}}', english_name)
                if '{{semester}}' in paragraph.text:
                    paragraph.text = paragraph.text.replace('{{semester}}', self.metadata.get('semester', '2026春'))
                if '{{date}}' in paragraph.text:
                    paragraph.text = paragraph.text.replace('{{date}}', self.metadata.get('date', ''))
            
            return doc
        except Exception as e:
            print(f"❌ 生成教学大纲失败: {e}")
            return None
    
    def generate_all_materials(self):
        """生成所有期初资料"""
        if not self.load_course_group_data():
            return False
        
        self.load_course_info()
        
        # 合并相同课程代码的课程信息
        merged_courses = {}
        for course in self.courses:
            course_code = course.get('courseCode', '')
            if course_code not in merged_courses:
                merged_courses[course_code] = course.copy()
            else:
                if 'teacherName' in course and course['teacherName']:
                    existing_teachers = merged_courses[course_code].get('teacherName', '')
                    if existing_teachers:
                        merged_courses[course_code]['teacherName'] = f"{existing_teachers}、{course['teacherName']}"
                    else:
                        merged_courses[course_code]['teacherName'] = course['teacherName']
        
        # 设置元数据
        self.metadata = {
            'date': datetime.now().strftime('%Y年%m月%d日'),
            'semester': '2026春'
        }
        
        # 为每门课程生成资料
        for course_code, course in merged_courses.items():
            course['formattedSemester'] = self.metadata['semester']
            
            # 创建课程文件夹（替换特殊字符）
            course_name = course.get('courseName', '')
            safe_course_name = course_name.replace('/', '_').replace('\\', '_').replace(':', '_').replace('*', '_').replace('?', '_').replace('"', '_').replace('<', '_').replace('>', '_').replace('|', '_')
            course_folder_name = f"{course_code}_{safe_course_name}"
            output_dir = os.path.join('智能金融学院_未知教师', course_folder_name)
            os.makedirs(output_dir, exist_ok=True)
            
            # 生成教学大纲
            doc = self.generate_syllabus(course)
            if doc:
                teacher_name = course.get('teacherName', '未知教师')
                safe_filename_name = course_name.replace('/', '_').replace('\\', '_').replace(':', '_').replace('*', '_').replace('?', '_').replace('"', '_').replace('<', '_').replace('>', '_').replace('|', '_')
                output_file = os.path.join(output_dir, f"2026春-{safe_filename_name}-教学大纲-{teacher_name}.docx")
                doc.save(output_file)
                print(f"✅ 教学大纲已生成: {output_file}")
            
            # 生成附件3
            wb3 = self.generate_attachment3(course)
            if wb3:
                course_leader = self.get_course_leader(course_code) or '未知教师'
                safe_filename_name = course_name.replace('/', '_').replace('\\', '_').replace(':', '_').replace('*', '_').replace('?', '_').replace('"', '_').replace('<', '_').replace('>', '_').replace('|', '_')
                output_file = os.path.join(output_dir, f"{safe_filename_name}-课程组期初教学资料检查情况记录表-{course_leader}.xlsx")
                wb3.save(output_file)
                print(f"✅ 附件3已生成: {output_file}")
            
            # 生成附件1
            wb1 = self.generate_attachment1(course)
            if wb1:
                department = course.get('department', '智能金融学院')
                course_leader = self.get_course_leader(course_code) or '未知教师'
                output_file = os.path.join(output_dir, f"{department}-教学大纲审核汇总表-{course_leader}.xlsx")
                wb1.save(output_file)
                print(f"✅ 附件1已生成: {output_file}")
        
        print("\n🎉 所有期初资料生成完成！")
        return True

if __name__ == "__main__":
    generator = SemesterMaterialsGenerator()
    generator.generate_all_materials()
