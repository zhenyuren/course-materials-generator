from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# 读取生成的文件
wb = load_workbook('智能金融学院_未知教师/FIN3004A_大数据分析基础/大数据分析基础-课程组期初教学资料检查情况记录表-任渝.xlsx')
ws = wb.active

print('最终格式验证V3:')
print('=' * 80)

# 检查表末尾端结构位置
print('表末尾端结构位置:')
total_cell = ws.cell(row=19, column=1)
print(f'合计行（第19行）: {total_cell.value}')

materials_cell = ws.cell(row=19, column=5)
print(f'资料份（第19行）: {materials_cell.value}')

# 总体评价意见只占一格（第19行第7列）
evaluation_cell = ws.cell(row=19, column=7)
print(f'总体评价意见（第19行第7列，只占一格）: {evaluation_cell.value}')

# 课程组负责人签字和日期在表右边部分（第7列），无边框
leader_cell = ws.cell(row=20, column=7)
print(f'课程组负责人签字（第20行第7列）: {leader_cell.value}')

date_cell = ws.cell(row=21, column=7)
print(f'日期（第21行第7列）: {date_cell.value}')

# 检查边框
print('\n边框检查:')
print(f'课程组负责人签字边框: {leader_cell.border}')
print(f'日期边框: {date_cell.border}')

# 检查合并单元格
print('\n合并单元格:')
merged_list = list(ws.merged_cells.ranges)
for merged_cell in merged_list:
    print(f'  {merged_cell}')

# 检查总体评价意见是否只占一格（不是合并单元格）
print('\n总体评价意见单元格检查:')
cell_address = f'G19'
is_merged = any(cell_address in str(merged) for merged in merged_list)
print(f'总体评价意见单元格G19是否为合并单元格: {is_merged}')

print('\n' + '=' * 80)
print('验证完成！')