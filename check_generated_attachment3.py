from openpyxl import load_workbook

# 查看生成的附件3文件
wb = load_workbook('智能金融学院_未知教师/FIN3004A_大数据分析基础/大数据分析基础-课程组期初教学资料检查情况记录表-任渝.xlsx')
ws = wb.active

print('附件3生成文件内容:')
print('=' * 80)
for row_idx in range(1, min(ws.max_row + 1, 15)):
    row_values = []
    for col_idx in range(1, min(ws.max_column + 1, 15)):
        cell_value = ws.cell(row=row_idx, column=col_idx).value
        row_values.append(str(cell_value) if cell_value is not None else 'None')
    print(f"行 {row_idx}: {', '.join(row_values)}")