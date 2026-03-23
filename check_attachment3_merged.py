from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# 查看生成的附件3文件
wb = load_workbook('智能金融学院_未知教师/FIN3004A_大数据分析基础/大数据分析基础-课程组期初教学资料检查情况记录表-任渝.xlsx')
ws = wb.active

print('附件3合并单元格详细检查:')
print('=' * 100)

# 基本信息
print(f'工作表名称: {ws.title}')
print(f'总行数: {ws.max_row}')
print(f'总列数: {ws.max_column}')

# 合并单元格检查
merged_list = list(ws.merged_cells.ranges)
print(f'\n合并单元格数量: {len(merged_list)}')
for i, merged_cell in enumerate(merged_list, 1):
    print(f'  {i}. {merged_cell}')

# 检查第2、3、4行的合并情况
print('\n第2、3、4行合并单元格检查:')
rows_to_check = [2, 3, 4]
cols_to_check = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']

for row_idx in rows_to_check:
    print(f'\n行 {row_idx}:')
    for col_letter in cols_to_check:
        cell_address = f'{col_letter}{row_idx}'
        is_merged = any(cell_address in str(mc) for mc in merged_list)
        print(f'  单元格 {cell_address}: {"✅ 已合并" if is_merged else "❌ 未合并"}')

# 检查表头格式
print('\n表头格式检查:')
headers = ['课程代码', '课程名', '开课单位', '使用年级/层次/专业', '归属校区', '教师（执笔人）', '课程组验收负责人']
for i, header in enumerate(headers, 2):  # 从B列开始
    cell = ws.cell(row=2, column=i)
    print(f'  单元格 {get_column_letter(i)}2: 值="{cell.value}", 字体={cell.font.name}, 大小={cell.font.size}, 加粗={cell.font.bold}')

# 检查评价指标列的背景色
print('\n评价指标列背景色检查:')
metric_cols = [9, 10, 11, 12, 13, 14]  # I-N列
for col_idx in metric_cols:
    cell = ws.cell(row=3, column=col_idx)
    fill = cell.fill
    has_gray_bg = fill.fgColor.index == '00DDDDDD'
    print(f'  单元格 {get_column_letter(col_idx)}3: {"✅ 浅灰色背景" if has_gray_bg else "❌ 无背景色"}')

# 检查分数列的背景色
print('\n分数列背景色检查:')
for col_idx in metric_cols:
    cell = ws.cell(row=4, column=col_idx)
    fill = cell.fill
    has_gray_bg = fill.fgColor.index == '00DDDDDD'
    print(f'  单元格 {get_column_letter(col_idx)}4: {"✅ 浅灰色背景" if has_gray_bg else "❌ 无背景色"}')

# 检查表末尾端结构
print('\n表末尾端结构检查:')

# 合计行
total_cell = ws.cell(row=19, column=1)
print(f'第19行合计: {"✅ 存在" if total_cell.value == "合计" else "❌ 缺失"}')

# 总体评价意见
evaluation_cell = ws.cell(row=19, column=7)
print(f'总体评价意见: {"✅ 存在" if evaluation_cell.value == "总体评价意见" else "❌ 缺失"}')

# 课程组负责人签字
leader_cell = ws.cell(row=20, column=1)
print(f'课程组负责人签字: {"✅ 存在" if "课程组负责人签字" in str(leader_cell.value) else "❌ 缺失"}')

# 日期行
date_cell = ws.cell(row=21, column=1)
print(f'日期行: {"✅ 存在" if "日期" in str(date_cell.value) else "❌ 缺失"}')

# 温馨提醒
note1_cell = ws.cell(row=23, column=1)
print(f'温馨提醒第1条: {"✅ 存在" if "温馨提醒" in str(note1_cell.value) else "❌ 缺失"}')

note2_cell = ws.cell(row=24, column=1)
print(f'温馨提醒第2条: {"✅ 存在" if "60-70分" in str(note2_cell.value) else "❌ 缺失"}')

note3_cell = ws.cell(row=25, column=1)
print(f'温馨提醒第3条: {"✅ 存在" if "评价栏应对本课程组" in str(note3_cell.value) else "❌ 缺失"}')

# 数据填充检查
print('\n数据填充检查:')
data_cells = [
    ('课程代码', 2, 5),
    ('课程名', 3, 5),
    ('开课单位', 4, 5),
    ('使用年级/层次/专业', 6, 5),
    ('教师（执笔人）', 7, 5),
    ('课程组验收负责人', 8, 5)
]
for label, col_idx, row_idx in data_cells:
    cell_value = ws.cell(row=row_idx, column=col_idx).value
    print(f'  {label}: {cell_value}')

print('\n' + '=' * 100)
print('检查完成！')