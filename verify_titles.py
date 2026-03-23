from openpyxl import load_workbook

print("验证标题设置：")
print("-" * 50)

# 验证附件3的标题
file_path3 = "智能金融学院_未知教师/FIN0008B_财经公文写作/财经公文写作-课程组期初教学资料检查情况记录表-未知教师.xlsx"
wb3 = load_workbook(file_path3)
ws3 = wb3.active

print("附件3 - 财经公文写作标题：")
print(f"{ws3.cell(row=1, column=1).value}")
print()

# 验证附件1的标题
file_path1 = "智能金融学院_未知教师/FIN0008B_财经公文写作/财经公文写作-教学大纲审核汇总表-未知教师.xlsx"
wb1 = load_workbook(file_path1)
ws1 = wb1.active

print("附件1 - 财经公文写作标题：")
print(f"{ws1.cell(row=1, column=1).value}")
print()

# 验证其他课程的标题
print("大数据分析基础 - 附件3标题：")
file_path4 = "智能金融学院_未知教师/FIN3004A_大数据分析基础/大数据分析基础-课程组期初教学资料检查情况记录表-未知教师.xlsx"
wb4 = load_workbook(file_path4)
ws4 = wb4.active
print(f"{ws4.cell(row=1, column=1).value}")
