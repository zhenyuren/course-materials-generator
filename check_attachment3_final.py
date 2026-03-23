from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# 查看生成的附件3文件
wb = load_workbook('智能金融学院_未知教师/FIN3004A_大数据分析基础/大数据分析基础-课程组期初教学资料检查情况记录表-任渝.xlsx')
ws = wb.active

print('附件3最终格式检查:')
print('=' * 100)

# 基本信息
print(f'工作表名称: {ws.title}')
print(f'总行数: {ws.max_row}')
print(f'总列数: {ws.max_column}')

# 合并单元格检查
merged_list = list(ws.merged_cells.ranges)
print(f'\n合并单元格数量: {len(merged_list)}')
for i, merged_cell in enumerate(merged_list, 1):
    print(f'  {i}. {merged_cell}')

# 检查A2:A4合并
a2_a4_merged = any(str(mc) == 'A2:A4' for mc in merged_list)
print(f'\nA2:A4合并单元格: {"✅ 已合并" if a2_a4_merged else "❌ 未合并"}')

# 标题检查
title_cell = ws.cell(row=1, column=1)
print(f'\n标题: {title_cell.value}')
print(f'标题字体: {title_cell.font.name}, 大小: {title_cell.font.size}, 加粗: {title_cell.font.bold}')

# 检查评价指标列的背景色
print('\n评价指标列背景色检查:')
metric_cols = [9, 10, 11, 12, 13, 14]  # I-N列
for col_idx in metric_cols:
    cell = ws.cell(row=3, column=col_idx)
    fill = cell.fill
    has_gray_bg = fill.fgColor.index == '00DDDDDD'
    print(f'  单元格 {get_column_letter(col_idx)}3: {"✅ 浅灰色背景" if has_gray_bg else "❌ 无背景色"}')

# 检查分数列的背景色
print('\n分数列背景色检查:')
for col_idx in metric_cols:
    cell = ws.cell(row=4, column=col_idx)
    fill = cell.fill
    has_gray_bg = fill.fgColor.index == '00DDDDDD'
    print(f'  单元格 {get_column_letter(col_idx)}4: {"✅ 浅灰色背景" if has_gray_bg else "❌ 无背景色"}')

# 检查表后面的内容
print('\n表后内容检查:')

# 合计行
total_cell = ws.cell(row=19, column=1)
print(f'第19行合计: {"✅ 存在" if total_cell.value == "合计" else "❌ 缺失"}')

# 总体评价意见
evaluation_cell = ws.cell(row=19, column=7)
print(f'总体评价意见: {"✅ 存在" if evaluation_cell.value == "总体评价意见" else "❌ 缺失"}')

# 课程组负责人签字
leader_cell = ws.cell(row=20, column=1)
print(f'课程组负责人签字: {"✅ 存在" if "课程组负责人签字" in str(leader_cell.value) else "❌ 缺失"}')

# 日期行
date_cell = ws.cell(row=21, column=1)
print(f'日期行: {"✅ 存在" if "日期" in str(date_cell.value) else "❌ 缺失"}')

# 温馨提醒
note1_cell = ws.cell(row=23, column=1)
print(f'温馨提醒第1条: {"✅ 存在" if "温馨提醒" in str(note1_cell.value) else "❌ 缺失"}')

note2_cell = ws.cell(row=24, column=1)
print(f'温馨提醒第2条: {"✅ 存在" if "60-70分" in str(note2_cell.value) else "❌ 缺失"}')

note3_cell = ws.cell(row=25, column=1)
print(f'温馨提醒第3条: {"✅ 存在" if "评价栏应对本课程组" in str(note3_cell.value) else "❌ 缺失"}')

# 数据填充检查
print('\n数据填充检查:')
data_cells = [
    ('课程代码', 2, 5),
    ('课程名', 3, 5),
    ('开课单位', 4, 5),
    ('使用年级/层次/专业', 6, 5),
    ('教师（执笔人）', 7, 5),
    ('课程组验收负责人', 8, 5)
]
for label, col_idx, row_idx in data_cells:
    cell_value = ws.cell(row=row_idx, column=col_idx).value
    print(f'  {label}: {cell_value}')

print('\n' + '=' * 100)
print('检查完成！')