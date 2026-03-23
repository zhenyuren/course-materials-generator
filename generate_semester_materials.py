import json
import os
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
import random

class SemesterMaterialsGenerator:
    def __init__(self, template_dir="期初资料"):
        self.template_dir = template_dir
        self.courses = []
        self.metadata = {}
        self.json_data = {}
    
    def load_course_group_data(self):
        """加载课程分组数据"""
        course_group_file = os.path.join(self.template_dir, "课程信息_2026年03月17日.json")
        if not os.path.exists(course_group_file):
            print(f"❌ 课程分组数据文件不存在: {course_group_file}")
            return False
        
        try:
            with open(course_group_file, 'r', encoding='utf-8') as f:
                data = json.load(f)
                # 检查是否包含courses字段
                if isinstance(data, dict) and 'courses' in data:
                    self.courses = data['courses']
                else:
                    self.courses = data
            print(f"✅ 成功加载 {len(self.courses)} 条课程分组数据")
            return True
        except Exception as e:
            print(f"❌ 加载课程分组数据失败: {e}")
            return False
    
    def load_json_data(self):
        """加载所有课程信息数据"""
        self.json_data = {}
        json_dir = "json"
        if not os.path.exists(json_dir):
            print(f"❌ JSON数据目录不存在: {json_dir}")
            return False
        
        for filename in os.listdir(json_dir):
            if filename.endswith('.json'):
                file_path = os.path.join(json_dir, filename)
                try:
                    with open(file_path, 'r', encoding='utf-8') as f:
                        data = json.load(f)
                        # 使用courseCode作为键
                        if 'courseCode' in data:
                            self.json_data[data['courseCode']] = data
                        else:
                            print(f"⚠️  文件 {filename} 缺少courseCode字段")
                except Exception as e:
                    print(f"❌ 加载 {filename} 失败: {e}")
        
        print(f"✅ 成功加载 {len(self.json_data)} 条课程信息数据")
        return True
    
    def get_course_leader(self, course_code):
        """根据课程代码获取课程负责人"""
        for course in self.courses:
            if course.get('courseCode') == course_code:
                return course.get('teacherName', '')
        return ''
    
    def generate_attachment3(self, course):
        """生成附件3：教学大纲审核表（为每门课程生成一个）"""
        import os
        
        # 创建新的工作簿
        wb = load_workbook(os.path.join(self.template_dir, "附件3 课程组教学资料检查情况记录表-完全一致模板.xlsx"))
        ws = wb.active
        
        # 填充课程信息
        course_code = course.get('courseCode', '')
        course_name = course.get('courseName', '')
        course_leader = self.get_course_leader(course_code)
        department = course.get('department', '')
        applicable_scope = course.get('applicableScope', '')
        teacher_name = course.get('teacherName', '')
        
        # 设置标题行（使用占位符转换方法）
        title_cell = ws.cell(row=1, column=1)
        title_cell.value = f"{course.get('formattedSemester', '2026春')}学期（{course.get('courseName', '')}）课程组期初教学资料检查情况记录表"
        
        # 在第5行开始填充数据
        ws.cell(row=5, column=1, value='1')
        ws.cell(row=5, column=2, value=course_code)
        ws.cell(row=5, column=3, value=course_name)
        ws.cell(row=5, column=4, value=department)
        
        # 解析applicableScope提取归属校区和使用年级/层次/专业
        campus = ""
        grade_major = applicable_scope
        
        # 检查是否包含具体区域信息
        if applicable_scope.startswith(('东区', '绵阳校区', '德阳校区')):
            if applicable_scope.startswith('东区'):
                campus = '东区'
                grade_major = applicable_scope[2:]
            elif applicable_scope.startswith('绵阳校区'):
                campus = '绵阳校区'
                grade_major = applicable_scope[5:]
            elif applicable_scope.startswith('德阳校区'):
                campus = '德阳校区'
                grade_major = applicable_scope[5:]
        else:
            # 默认西区
            campus = '西区'
            grade_major = applicable_scope
        
        ws.cell(row=5, column=5, value=grade_major)
        ws.cell(row=5, column=6, value=campus)
        ws.cell(row=5, column=7, value=teacher_name)
        ws.cell(row=5, column=8, value=course_leader)
        
        # 生成分数，满足总分86-92之间
        while True:
            goal_score = random.randint(12, 14)
            norm_score = random.randint(8, 9)
            method_score = random.randint(8, 9)
            schedule_score = random.randint(12, 14)
            design_score = random.randint(42, 48)
            
            total = goal_score + norm_score + method_score + schedule_score + design_score
            
            if 86 <= total <= 92:
                break
        
        # 填充分数
        ws.cell(row=5, column=9, value=goal_score)
        ws.cell(row=5, column=10, value=norm_score)
        ws.cell(row=5, column=11, value=method_score)
        ws.cell(row=5, column=12, value=schedule_score)
        ws.cell(row=5, column=13, value=design_score)
        ws.cell(row=5, column=14, value=total)
        
        # 更新日期
        current_date = self.metadata.get('date', '')
        if current_date:
            # 设置日期（第20行第8列开始的合并区域）
            date_cell = ws.cell(row=20, column=8)
            date_cell.value = f"日期：{current_date}"
        
        # 添加电子签名图片
        signature_image_path = os.path.join(self.template_dir, "任渝.png")
        if os.path.exists(signature_image_path):
            try:
                img = Image(signature_image_path)
                img.width = 100
                img.height = 40
                ws.add_image(img, 'J19')
                print(f"✅ 电子签名已添加到 {course_name} 的附件3")
            except Exception as e:
                print(f"⚠️ 添加电子签名失败: {e}")
        
        return wb
    
    def generate_attachment1(self, course):
        """生成附件1：教学大纲审核汇总表（为每门课程生成一个）"""
        import os
        
        # 使用模板文件
        template_file = os.path.join(self.template_dir, "附件1 教学大纲审核汇总表-完全一致模板.xlsx")
        if not os.path.exists(template_file):
            print(f"❌ 附件1模板文件不存在: {template_file}")
            return None
        
        try:
            # 加载工作簿，保持原始格式
            wb = load_workbook(template_file)
            ws = wb.active
            
            # 填充课程信息
            course_code = course.get('courseCode', '')
            course_name = course.get('courseName', '')
            course_leader = self.get_course_leader(course_code)
            department = course.get('department', '')
            applicable_scope = course.get('applicableScope', '')
            teacher_name = course.get('teacherName', '')
            
            # 设置标题行（使用占位符转换方法）
            title_cell = ws.cell(row=1, column=1)
            title_cell.value = f"{course.get('formattedSemester', '2026春')}学期（{course.get('courseName', '')}）教学大纲审核汇总表"
            
            # 解析applicableScope提取归属校区和使用年级/层次/专业
            campus = ""
            grade_major = applicable_scope
            
            # 检查是否包含具体区域信息
            if applicable_scope.startswith(('东区', '绵阳校区', '德阳校区')):
                if applicable_scope.startswith('东区'):
                    campus = '东区'
                    grade_major = applicable_scope[2:]
                elif applicable_scope.startswith('绵阳校区'):
                    campus = '绵阳校区'
                    grade_major = applicable_scope[5:]
                elif applicable_scope.startswith('德阳校区'):
                    campus = '德阳校区'
                    grade_major = applicable_scope[5:]
            else:
                # 默认西区
                campus = '西区'
                grade_major = applicable_scope
            
            # 在第5行开始填充数据
            ws.cell(row=5, column=1, value='1')  # 序号
            ws.cell(row=5, column=2, value=course_code)  # 课程代码
            ws.cell(row=5, column=3, value=course_name)  # 课程名称
            ws.cell(row=5, column=4, value=department)  # 开课单位
            ws.cell(row=5, column=5, value=campus)  # 所属校区
            ws.cell(row=5, column=6, value=grade_major)  # 使用年级/层次/专业
            ws.cell(row=5, column=7, value=teacher_name)  # 执笔人
            ws.cell(row=5, column=8, value=course_leader)  # 验收负责人
            
            # 生成分数，满足总分86-92之间，且单项分数小于对应分数
            import random
            while True:
                goal_score = random.randint(12, 14)  # 人才培养目标符合度（满分15）
                norm_score = random.randint(8, 9)     # 大纲撰写规范性（满分10）
                method_score = random.randint(8, 9)   # 教学方式（满分10）
                schedule_score = random.randint(12, 14)# 教学进度及安排（满分15）
                design_score = random.randint(42, 48) # 课程设计和框架（满分50）
                
                total = goal_score + norm_score + method_score + schedule_score + design_score
                
                if 86 <= total <= 92:
                    break
            
            # 填充分数
            ws.cell(row=5, column=9, value=goal_score)     # 人才培养目标符合度
            ws.cell(row=5, column=10, value=norm_score)   # 大纲撰写规范性
            ws.cell(row=5, column=11, value=method_score) # 教学方式
            ws.cell(row=5, column=12, value=schedule_score) # 教学进度及安排
            ws.cell(row=5, column=13, value=design_score) # 课程设计和框架
            ws.cell(row=5, column=14, value=total)       # 总分
            
            # 计算资料份数
            teacher_count = 0
            for c in self.courses:
                if c.get('courseCode') == course_code:
                    teacher_count += 1
            
            # 更新资料份数（第18行第1列开始的合并区域）
            materials_cell = ws.cell(row=18, column=1)
            materials_cell.value = f'资料 {teacher_count} 份'
            
            # 更新日期
            current_date = self.metadata.get('date', '')
            if current_date:
                # 设置日期（第20行第10列开始的合并区域）
                date_cell = ws.cell(row=20, column=10)
                date_cell.value = f"日期：{current_date}"
            
            # 添加电子签名图片（在开课单位负责人签字后面）
            from openpyxl.drawing.image import Image
            
            # 设置图片位置（开课单位负责人签字后面，第19行第10列）
            img_position = 'J19'
            
            # 检查是否有电子签名图片（从期初资料文件夹读取）
            signature_image_path = os.path.join(self.template_dir, "任渝.png")
            
            if os.path.exists(signature_image_path):
                try:
                    # 创建图片对象
                    img = Image(signature_image_path)
                    
                    # 设置图片大小（根据需要调整）
                    img.width = 100
                    img.height = 40
                    
                    # 将图片添加到指定位置
                    ws.add_image(img, img_position)
                    print(f"✅ 电子签名已添加到 {course_name} 的附件1")
                except Exception as e:
                    print(f"⚠️ 添加电子签名失败: {e}")
            else:
                print(f"⚠️ 电子签名图片不存在，请将签名图片保存为 任渝.png 放在期初资料文件夹")
            
            return wb
        except Exception as e:
            print(f"❌ 生成附件1失败: {e}")
            return None
    
    def generate_all_materials(self):
        """生成所有期初资料"""
        # 加载期初资料的课程分组数据
        if not self.load_course_group_data():
            return False
        
        # 加载任渝的课程信息
        renyu_course_file = os.path.join("json", "任渝_课程信息_2026年03月21日.json")
        if os.path.exists(renyu_course_file):
            try:
                with open(renyu_course_file, 'r', encoding='utf-8') as f:
                    renyu_data = json.load(f)
                    if isinstance(renyu_data, dict) and 'courses' in renyu_data:
                        self.courses.extend(renyu_data['courses'])
                    else:
                        self.courses.extend(renyu_data)
                print(f"✅ 成功加载任渝的课程信息")
            except Exception as e:
                print(f"❌ 加载任渝课程信息失败: {e}")
        
        # 合并相同课程代码的课程信息
        merged_courses = {}
        for course in self.courses:
            course_code = course.get('courseCode', '')
            if course_code not in merged_courses:
                merged_courses[course_code] = course.copy()
            else:
                # 合并教师信息
                if 'teacherName' in course and course['teacherName']:
                    existing_teachers = merged_courses[course_code].get('teacherName', '')
                    if existing_teachers:
                        merged_courses[course_code]['teacherName'] = f"{existing_teachers}、{course['teacherName']}"
                    else:
                        merged_courses[course_code]['teacherName'] = course['teacherName']
        
        # 设置元数据
        self.metadata = {
            'date': datetime.now().strftime('%Y年%m月%d日'),
            'semester': '2026春'
        }
        
        # 为每门课程生成资料
        for course_code, course in merged_courses.items():
            course['formattedSemester'] = self.metadata['semester']
            
            # 生成附件1
            wb1 = self.generate_attachment1(course)
            if wb1:
                # 创建课程文件夹
                course_folder_name = f"{course_code}_{course.get('courseName', '')}"
                output_dir = os.path.join('智能金融学院_未知教师', course_folder_name)
                os.makedirs(output_dir, exist_ok=True)
                output_file = os.path.join(output_dir, f"{course.get('courseName', '')}-教学大纲审核汇总表-未知教师.xlsx")
                wb1.save(output_file)
                print(f"✅ 附件1已生成: {output_file}")
            
            # 生成附件3
            wb3 = self.generate_attachment3(course)
            if wb3:
                # 创建课程文件夹
                course_folder_name = f"{course_code}_{course.get('courseName', '')}"
                output_dir = os.path.join('智能金融学院_未知教师', course_folder_name)
                os.makedirs(output_dir, exist_ok=True)
                output_file = os.path.join(output_dir, f"{course.get('courseName', '')}-课程组期初教学资料检查情况记录表-未知教师.xlsx")
                wb3.save(output_file)
                print(f"✅ 附件3已生成: {output_file}")
        
        print("\n🎉 所有期初资料生成完成！")
        return True

if __name__ == "__main__":
    generator = SemesterMaterialsGenerator()
    generator.generate_all_materials()