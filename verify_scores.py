from openpyxl import load_workbook
import json

# 读取JSON数据
with open('json/任渝_课程信息_2026年03月21日.json', 'r', encoding='utf-8') as f:
    data = json.load(f)
    courses = data.get('courses', [])

print('分数验证:')
print('=' * 80)

# 验证每门课程的分数
for course in courses:
    course_code = course.get('courseCode', '')
    course_name = course.get('courseName', '')
    
    # 跳过重复课程
    if course_code == 'FIN4011B' and course_name == '跨境电商大数据统计与分析':
        continue
    if course_code == 'FIN3005A' and course_name == '大数据处理技术':
        continue
    
    print(f'\n课程: {course_name} ({course_code})')
    
    # 读取生成的文件
    file_path = f'智能金融学院_未知教师/{course_code}_{course_name}/{course_name}-课程组期初教学资料检查情况记录表-任渝.xlsx'
    try:
        wb = load_workbook(file_path)
        ws = wb.active
        
        # 获取分数
        outline_score = ws.cell(row=5, column=9).value
        norm_score = ws.cell(row=5, column=10).value
        objective_score = ws.cell(row=5, column=11).value
        method_score = ws.cell(row=5, column=12).value
        process_score = ws.cell(row=5, column=13).value
        total_score = ws.cell(row=5, column=14).value
        
        print(f'教学大纲符合度: {outline_score} (满分15)')
        print(f'撰写规范性: {norm_score} (满分10)')
        print(f'教学目标设计: {objective_score} (满分15)')
        print(f'教学方法设计: {method_score} (满分10)')
        print(f'教学过程设计: {process_score} (满分50)')
        print(f'总分: {total_score}')
        
        # 验证分数
        valid = True
        
        if outline_score >= 15:
            print('❌ 教学大纲符合度分数必须小于15')
            valid = False
        
        if norm_score >= 10:
            print('❌ 撰写规范性分数必须小于10')
            valid = False
        
        if objective_score >= 15:
            print('❌ 教学目标设计分数必须小于15')
            valid = False
        
        if method_score >= 10:
            print('❌ 教学方法设计分数必须小于10')
            valid = False
        
        if process_score >= 50:
            print('❌ 教学过程设计分数必须小于50')
            valid = False
        
        if total_score < 86 or total_score > 92:
            print('❌ 总分必须在86-92之间')
            valid = False
        
        if valid:
            print('✅ 分数验证通过！')
            
        # 验证课程组负责人签字和日期位置
        leader_cell = ws.cell(row=20, column=14).value
        date_cell = ws.cell(row=21, column=12).value
        
        print(f'\n课程组负责人签字位置: 第20行第14列 (N列): {leader_cell}')
        print(f'日期位置: 第21行第12列 (L列): {date_cell}')
        
    except FileNotFoundError:
        print(f'❌ 文件不存在: {file_path}')

print('\n' + '=' * 80)
print('验证完成！')