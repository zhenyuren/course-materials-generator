from openpyxl import load_workbook
import json

# 读取JSON数据
with open('json/任渝_课程信息_2026年03月21日.json', 'r', encoding='utf-8') as f:
    data = json.load(f)
    courses = data.get('courses', [])

print('数据提取验证:')
print('=' * 80)

# 验证每门课程的数据提取
for course in courses:
    course_code = course.get('courseCode', '')
    course_name = course.get('courseName', '')
    applicable_scope = course.get('applicableScope', '')
    
    print(f'\n课程: {course_name} ({course_code})')
    print(f'原始applicableScope: {applicable_scope}')
    
    # 解析applicableScope提取归属校区和使用年级/层次/专业
    campus = ""
    grade_major = applicable_scope
    
    if applicable_scope.startswith(('东区', '西区', '南区', '北区')):
        if applicable_scope.startswith('东区'):
            campus = '东区'
            grade_major = applicable_scope[2:]
        elif applicable_scope.startswith('西区'):
            campus = '西区'
            grade_major = applicable_scope[2:]
        elif applicable_scope.startswith('南区'):
            campus = '南区'
            grade_major = applicable_scope[2:]
        elif applicable_scope.startswith('北区'):
            campus = '北区'
            grade_major = applicable_scope[2:]
    
    print(f'提取的归属校区: {campus}')
    print(f'提取的使用年级/层次/专业: {grade_major}')

# 验证生成的Excel文件
print('\n' + '=' * 80)
print('生成文件数据验证:')
print('=' * 80)

# 读取生成的文件
wb = load_workbook('智能金融学院_未知教师/FIN3004A_大数据分析基础/大数据分析基础-课程组期初教学资料检查情况记录表-任渝.xlsx')
ws = wb.active

print('\n大数据分析基础课程数据:')
print(f'课程代码: {ws.cell(row=5, column=2).value}')
print(f'课程名: {ws.cell(row=5, column=3).value}')
print(f'开课单位: {ws.cell(row=5, column=4).value}')
print(f'使用年级/层次/专业: {ws.cell(row=5, column=5).value}')
print(f'归属校区: {ws.cell(row=5, column=6).value}')
print(f'教师（执笔人）: {ws.cell(row=5, column=7).value}')
print(f'课程组验收负责人: {ws.cell(row=5, column=8).value}')

print('\n' + '=' * 80)
print('验证完成！')