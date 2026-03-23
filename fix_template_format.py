from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

# 加载当前模板
wb = load_workbook('期初资料/附件1 教学大纲的审核汇总表-大数据信息分析_任渝.xlsx')
ws = wb.active

print('当前模板文件信息:')
print(f'工作表名称: {ws.title}')
print(f'总行数: {ws.max_row}')
print(f'总列数: {ws.max_column}')
print(f'合并单元格数量: {len(list(ws.merged_cells.ranges))}')

# 检查前5行的内容和格式
print('\n前5行内容:')
for row_idx in range(1, 6):
    row_data = []
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell_value = cell.value
        row_data.append(str(cell_value) if cell_value is not None else 'None')
    print(f'行 {row_idx}: {", ".join(row_data)}')

# 检查第一行的字体和对齐方式
print('\n第一行格式信息:')
for col_idx in range(1, 5):
    cell = ws.cell(row=1, column=col_idx)
    font = cell.font
    alignment = cell.alignment
    print(f'单元格 {get_column_letter(col_idx)}1:')
    print(f'  字体: {font.name}, 大小: {font.size}, 加粗: {font.bold}')
    print(f'  对齐方式: 水平={alignment.horizontal}, 垂直={alignment.vertical}')

# 创建一个新的模板文件，模拟第二张图的格式
new_wb = Workbook()
new_ws = new_wb.active
new_ws.title = "Sheet1"

# 设置列宽
column_widths = [10, 12, 15, 12, 10, 20, 10, 10, 12, 10, 10, 12, 12, 10]
for i, width in enumerate(column_widths, 1):
    new_ws.column_dimensions[get_column_letter(i)].width = width

# 设置标题行（合并单元格）
new_ws.merge_cells('A1:N1')
title_cell = new_ws['A1']
title_cell.value = '{{formattedSemester}}学期各开课单位教学大纲的审核汇总表'
title_cell.font = Font(name='宋体', size=16, bold=True)
title_cell.alignment = Alignment(horizontal='center', vertical='center')

# 设置表头
headers = ['序号', '课程代码', '课程名', '开课单位', '所属校区', '使用年级/层次/专业', '执笔人', '验收负责人', '教学大纲评价指标', '', '', '', '', '']
for i, header in enumerate(headers, 1):
    cell = new_ws.cell(row=2, column=i, value=header)
    cell.font = Font(name='宋体', size=12, bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center')

# 设置评价指标
metrics = ['', '', '', '', '', '', '', '', '人才培养目标符合度', '大纲撰写规范性', '教学方式', '教学进度及安排', '课程设计和框架', '总分']
for i, metric in enumerate(metrics, 1):
    cell = new_ws.cell(row=3, column=i, value=metric)
    cell.font = Font(name='宋体', size=11)
    cell.alignment = Alignment(horizontal='center', vertical='center')

# 设置分数
scores = ['', '', '', '', '', '', '', '', '15分', '10分', '10分', '15分', '50分', '100分']
for i, score in enumerate(scores, 1):
    cell = new_ws.cell(row=4, column=i, value=score)
    cell.font = Font(name='宋体', size=11)
    cell.alignment = Alignment(horizontal='center', vertical='center')

# 设置数据行
data = ['1', '{{courseCode}}', '{{courseName}}', '{{department}}', '绵阳', '{{applicableScope}}', '{{teacherName}}', '{{courseLeader}}', '12', '8', '8', '14', '43', '85']
for i, value in enumerate(data, 1):
    cell = new_ws.cell(row=5, column=i, value=value)
    cell.font = Font(name='宋体', size=11)
    cell.alignment = Alignment(horizontal='center', vertical='center')

# 设置行高
new_ws.row_dimensions[1].height = 30
new_ws.row_dimensions[2].height = 20
new_ws.row_dimensions[3].height = 18
new_ws.row_dimensions[4].height = 18
new_ws.row_dimensions[5].height = 18

# 保存新模板
new_wb.save('期初资料/附件1 教学大纲的审核汇总表-模板_new.xlsx')
print('\n✅ 新模板文件已创建: 期初资料/附件1 教学大纲的审核汇总表-模板_new.xlsx')