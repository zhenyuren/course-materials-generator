from openpyxl import load_workbook

# 加载生成的文件
generated_file = "智能金融学院_未知教师/RRD1324A_合作组织专业见习/合作组织专业见习-教学大纲审核汇总表-未知教师.xlsx"
wb = load_workbook(generated_file)
ws = wb.active

print("生成文件的格式检查：")
print(f"工作表名称: {ws.title}")
print(f"行数: {ws.max_row}, 列数: {ws.max_column}")

# 检查合并单元格
merged_cells = list(ws.merged_cells.ranges)
print(f"\n合并单元格数量: {len(merged_cells)}")
for merged_range in merged_cells:
    print(f"  {merged_range}")

# 检查标题行
title_cell = ws.cell(row=1, column=1)
print(f"\n标题内容: {title_cell.value}")
print(f"标题字体: {title_cell.font.name}, 大小: {title_cell.font.size}, 加粗: {title_cell.font.bold}")

# 检查表头行
print("\n表头行内容:")
for col in range(1, 15):
    cell = ws.cell(row=2, column=col)
    if cell.value:
        print(f"  第2行第{col}列: {cell.value}")

# 检查数据行
print("\n数据行内容:")
for col in range(1, 15):
    cell = ws.cell(row=5, column=col)
    if cell.value:
        print(f"  第5行第{col}列: {cell.value}")

# 检查合计行
print("\n合计行内容:")
total_cell = ws.cell(row=18, column=1)
materials_cell = ws.cell(row=18, column=1)
print(f"合计内容: {total_cell.value}")
print(f"资料份数: {materials_cell.value}")

# 检查日期行
date_cell = ws.cell(row=20, column=10)
print(f"\n日期内容: {date_cell.value}")

print("\n✅ 格式检查完成！")
