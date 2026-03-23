from openpyxl import load_workbook

# 加载模板文件
template_file = "期初资料/附件1 教学大纲审核汇总表-完全一致模板.xlsx"
wb = load_workbook(template_file)
ws = wb.active

# 检查第5行的数据填充区域（第1-14列）
print("第5行数据填充区域检查：")
for col in range(1, 15):
    cell = ws.cell(row=5, column=col)
    print(f"第5行第{col}列: {cell.value}")

# 检查第18行第2列（资料份数）
print("\n第18行第2列: ", ws.cell(row=18, column=2).value)

# 检查第20行第10列（日期）
print("第20行第10列: ", ws.cell(row=20, column=10).value)
