from openpyxl import load_workbook

# 加载模板文件
template_file = "期初资料/附件3 课程组教学资料检查情况记录表-完全一致模板.xlsx"
wb = load_workbook(template_file)
ws = wb.active

# 获取所有合并单元格
merged_cells = list(ws.merged_cells.ranges)

print("合并单元格列表：")
for merged_range in merged_cells:
    print(f"{merged_range}")

# 检查第20行第9列是否是合并单元格
cell = ws.cell(row=20, column=9)
print(f"\n第20行第9列是否是合并单元格: {cell.is_merged}")

# 如果是合并单元格，查看合并区域
if cell.is_merged:
    for merged_range in merged_cells:
        if merged_range.min_row <= 20 <= merged_range.max_row and merged_range.min_col <= 9 <= merged_range.max_col:
            print(f"合并区域: {merged_range}")
            print(f"左上角单元格: ({merged_range.min_row}, {merged_range.min_col})")
            print(f"右下角单元格: ({merged_range.max_row}, {merged_range.max_col})")
