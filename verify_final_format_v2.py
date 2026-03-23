from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# 读取生成的文件
wb = load_workbook('智能金融学院_未知教师/FIN3004A_大数据分析基础/大数据分析基础-课程组期初教学资料检查情况记录表-任渝.xlsx')
ws = wb.active

print('最终格式验证V2:')
print('=' * 80)

# 打印标题行
title = ws.cell(row=1, column=1).value
print(f'标题: {title}')

# 打印表头（第2行）
print('\n表头（第2行）:')
headers = []
for col_idx in range(1, 15):
    cell_value = ws.cell(row=2, column=col_idx).value
    headers.append(str(cell_value) if cell_value is not None else '')
print(' | '.join(headers))

# 打印评价指标行（第3行）
print('\n评价指标（第3行）:')
metrics = []
for col_idx in range(1, 15):
    cell_value = ws.cell(row=3, column=col_idx).value
    metrics.append(str(cell_value) if cell_value is not None else '')
print(' | '.join(metrics))

# 打印分数行（第4行）
print('\n分数（第4行）:')
scores = []
for col_idx in range(1, 15):
    cell_value = ws.cell(row=4, column=col_idx).value
    scores.append(str(cell_value) if cell_value is not None else '')
print(' | '.join(scores))

# 打印数据行（第5行）
print('\n数据行（第5行）:')
data = []
for col_idx in range(1, 15):
    cell_value = ws.cell(row=5, column=col_idx).value
    data.append(str(cell_value) if cell_value is not None else '')
print(' | '.join(data))

# 检查合并单元格
print('\n合并单元格:')
merged_list = list(ws.merged_cells.ranges)
for merged_cell in merged_list:
    print(f'  {merged_cell}')

# 检查格式
print('\n关键格式检查:')

# 标题格式
title_cell = ws.cell(row=1, column=1)
print(f'标题字体: {title_cell.font.name}, 大小: {title_cell.font.size}, 加粗: {title_cell.font.bold}')

# 表头格式（带换行和灰色背景）
print('\n表头格式:')
for col_idx in range(1, 9):
    cell = ws.cell(row=2, column=col_idx)
    print(f'表头列{col_idx}: 字体={cell.font.name}, 大小={cell.font.size}, 加粗={cell.font.bold}, 背景色={cell.fill.fgColor.index}, 换行={cell.alignment.wrap_text}')

# 评价指标列格式（字体小一些，灰色背景）
print('\n评价指标列格式:')
for col_idx in range(9, 15):
    cell = ws.cell(row=3, column=col_idx)
    print(f'评价指标列{col_idx}: 字体={cell.font.name}, 大小={cell.font.size}, 背景色={cell.fill.fgColor.index}')

# 分数列格式（字体小一些，灰色背景）
print('\n分数列格式:')
for col_idx in range(9, 15):
    cell = ws.cell(row=4, column=col_idx)
    print(f'分数列{col_idx}: 字体={cell.font.name}, 大小={cell.font.size}, 背景色={cell.fill.fgColor.index}')

# 检查表末尾端结构位置
print('\n表末尾端结构位置:')
total_cell = ws.cell(row=19, column=1)
print(f'合计行（第19行）: {total_cell.value}')

materials_cell = ws.cell(row=19, column=5)
print(f'资料份（第19行）: {materials_cell.value}')

evaluation_cell = ws.cell(row=19, column=7)
print(f'总体评价意见（第19行）: {evaluation_cell.value}')

# 课程组负责人签字和日期在表右边部分（I-N列）
leader_cell = ws.cell(row=20, column=9)
print(f'课程组负责人签字（第20行第9列）: {leader_cell.value}')

date_cell = ws.cell(row=21, column=9)
print(f'日期（第21行第9列）: {date_cell.value}')

# 检查边框
print('\n边框检查:')
print(f'课程组负责人签字边框: {leader_cell.border}')
print(f'日期边框: {date_cell.border}')

note1_cell = ws.cell(row=23, column=1)
print(f'温馨提醒1（第23行）: {note1_cell.value}')

note2_cell = ws.cell(row=24, column=1)
print(f'温馨提醒2（第24行）: {note2_cell.value}')

note3_cell = ws.cell(row=25, column=1)
print(f'温馨提醒3（第25行）: {note3_cell.value}')

# 检查数据填充
print('\n数据填充检查:')
print(f'课程代码: {ws.cell(row=5, column=2).value}')
print(f'课程名: {ws.cell(row=5, column=3).value}')
print(f'开课单位: {ws.cell(row=5, column=4).value}')
print(f'使用年级/层次/专业: {ws.cell(row=5, column=5).value}')
print(f'教师（执笔人）: {ws.cell(row=5, column=7).value}')
print(f'课程组验收负责人: {ws.cell(row=5, column=8).value}')

# 检查列宽
print('\n列宽检查:')
for col_idx in range(1, 15):
    col_letter = get_column_letter(col_idx)
    width = ws.column_dimensions[col_letter].width
    print(f'列{col_idx}({col_letter}): 宽度={width}')

print('\n' + '=' * 80)
print('验证完成！')