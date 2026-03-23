from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
import os

# 创建工作簿
wb = Workbook()
ws = wb.active

# 创建边框样式
thin_border = Border(
    left=Side(border_style='thin', color='000000'),
    right=Side(border_style='thin', color='000000'),
    top=Side(border_style='thin', color='000000'),
    bottom=Side(border_style='thin', color='000000')
)

# 创建无边框样式
no_border = Border(
    left=None,
    right=None,
    top=None,
    bottom=None
)

# 创建浅灰色背景
light_gray_fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')

# 设置列宽（根据用户截图调整）
column_widths = [5, 12, 20, 12, 22, 10, 15, 15, 12, 12, 12, 12, 15, 8]
for i, width in enumerate(column_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = width

# 设置标题行（合并A1:N1）
title_cell = ws['A1']
title_cell.value = '{{formattedSemester}}学期（{{courseName}}）课程组期初教学资料检查情况记录表'
title_cell.font = Font(name='宋体', size=16, bold=True)
title_cell.alignment = Alignment(horizontal='center', vertical='center')
title_cell.border = thin_border
ws.merge_cells('A1:N1')

# 设置表头（第2行），带换行和灰色背景
headers = [
    '序号',
    '课程代码',
    '课程名',
    '开课单位',
    '使用年级/层次/专业',
    '归属校区',
    '教师（执笔人）',
    '课程组验收负责人'
]
for i, header in enumerate(headers, 1):
    cell = ws.cell(row=2, column=i, value=header)
    cell.font = Font(name='宋体', size=12, bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border
    cell.fill = light_gray_fill

# 设置教案评价指标标题（第2行第9列），灰色背景
i2_cell = ws.cell(row=2, column=9, value='教案评价指标')
i2_cell.font = Font(name='宋体', size=12, bold=True)
i2_cell.alignment = Alignment(horizontal='center', vertical='center')
i2_cell.border = thin_border
i2_cell.fill = light_gray_fill

# 设置评价指标行（第3行），字体小一些
metrics = ['', '', '', '', '', '', '', '', '教学大纲符合度', '撰写规范性', '教学目标设计', '教学方法设计', '教学过程设计', '总分']
for i, metric in enumerate(metrics, 1):
    cell = ws.cell(row=3, column=i, value=metric)
    cell.font = Font(name='宋体', size=10)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border
    # 添加浅灰色背景
    if i >= 9:
        cell.fill = light_gray_fill

# 设置分数行（第4行），字体小一些
scores = ['', '', '', '', '', '', '', '', '15分', '10分', '15分', '10分', '50分', '100分']
for i, score in enumerate(scores, 1):
    cell = ws.cell(row=4, column=i, value=score)
    cell.font = Font(name='宋体', size=10)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border
    # 添加浅灰色背景
    if i >= 9:
        cell.fill = light_gray_fill

# 设置数据行占位符（第5行）
data = ['1', '{{courseCode}}', '{{courseName}}', '{{department}}', '{{gradeMajor}}', '{{campus}}', '{{teacherName}}', '', '', '', '', '', '', '']
for i, value in enumerate(data, 1):
    cell = ws.cell(row=5, column=i, value=value)
    cell.font = Font(name='宋体', size=11)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border

# 设置空行（第6-18行）
for row_idx in range(6, 19):
    for col_idx in range(1, 15):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.font = Font(name='宋体', size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

# 添加合计行（第19行）
total_cell = ws.cell(row=19, column=1, value='合计')
total_cell.font = Font(name='宋体', size=11, bold=True)
total_cell.alignment = Alignment(horizontal='center', vertical='center')
total_cell.border = thin_border

materials_cell = ws.cell(row=19, column=5, value='资料 {{materialsCount}} 份')
materials_cell.font = Font(name='宋体', size=11)
materials_cell.alignment = Alignment(horizontal='center', vertical='center')
materials_cell.border = thin_border

evaluation_cell = ws.cell(row=19, column=7, value='总体评价意见')
evaluation_cell.font = Font(name='宋体', size=11)
evaluation_cell.alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)
evaluation_cell.border = thin_border

# 更新总体评价意见内容
h19_cell = ws.cell(row=19, column=8, value='满足教学大纲要求，课堂时间分配合理，教学实施设计上有的放矢，教学重点、难点把握得当，课程组审核合格。')
h19_cell.font = Font(name='宋体', size=11)
h19_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
h19_cell.border = thin_border

# 添加课程组负责人签字行（第20行）- 只保留"课程组负责人签字："文字，签名图片将在生成时添加
leader_cell = ws.cell(row=20, column=7, value='课程组负责人签字：')
leader_cell.font = Font(name='宋体', size=11)
leader_cell.alignment = Alignment(horizontal='left', vertical='center')
leader_cell.border = no_border

# 添加日期行（第21行）- 在I列，JK合并显示日期
date_label_cell = ws.cell(row=21, column=7, value='日期：')
date_label_cell.font = Font(name='宋体', size=11)
date_label_cell.alignment = Alignment(horizontal='left', vertical='center')
date_label_cell.border = no_border

date_cell = ws.cell(row=21, column=9, value='{{currentDate}}')
date_cell.font = Font(name='宋体', size=11)
date_cell.alignment = Alignment(horizontal='left', vertical='center')
date_cell.border = no_border

# 添加温馨提醒（第23-25行）
note1_cell = ws.cell(row=23, column=1, value='温馨提醒：1.此表一式两份，开课单位保存一份，教务处保存一份')
note1_cell.font = Font(name='宋体', size=11)
note1_cell.alignment = Alignment(horizontal='left', vertical='center')

note2_cell = ws.cell(row=24, column=1, value='2.60-70分为基本合格；70-80分为合格；80-90分为良好；90分以上为优秀')
note2_cell.font = Font(name='宋体', size=11)
note2_cell.alignment = Alignment(horizontal='left', vertical='center')

note3_cell = ws.cell(row=25, column=1, value='3.评价栏应对本课程组教学资料编写情况进行描述（50字以上）')
note3_cell.font = Font(name='宋体', size=11)
note3_cell.alignment = Alignment(horizontal='left', vertical='center')

# 现在进行合并单元格操作
# 合并第2、3、4行的A-H列（这些列有灰色背景）
columns_to_merge = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']
for col_letter in columns_to_merge:
    ws.merge_cells(f'{col_letter}2:{col_letter}4')

# 合并第2行I-N列（教案评价指标，灰色背景）
ws.merge_cells('I2:N2')

# 合并合计行
ws.merge_cells('A19:D19')
ws.merge_cells('E19:F19')
ws.merge_cells('H19:N19')

# 合并日期后面的两个格子（J-K列）
ws.merge_cells('J21:K21')

# 合并温馨提醒行
ws.merge_cells('A23:N23')
ws.merge_cells('A24:N24')
ws.merge_cells('A25:N25')

# 保存文件
wb.save('期初资料/附件3 课程组教学资料检查情况记录表-完全一致模板.xlsx')
print('✅ 完全一致模板V11已创建: 期初资料/附件3 课程组教学资料检查情况记录表-完全一致模板.xlsx')