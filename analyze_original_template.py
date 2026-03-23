from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment

# 加载原始附件1模板
wb = load_workbook('期初资料/附件1 教学大纲的审核汇总表-大数据信息分析_任渝.xlsx')
ws = wb.active

print('原始附件1模板详细分析:')
print('=' * 100)

# 基本信息
print(f'工作表名称: {ws.title}')
print(f'总行数: {ws.max_row}')
print(f'总列数: {ws.max_column}')

# 合并单元格
merged_list = list(ws.merged_cells.ranges)
print(f'\n合并单元格数量: {len(merged_list)}')
for i, merged_cell in enumerate(merged_list, 1):
    print(f'  {i}. {merged_cell}')

# 列宽
print('\n列宽信息:')
for col_idx in range(1, ws.max_column + 1):
    col_letter = get_column_letter(col_idx)
    width = ws.column_dimensions[col_letter].width
    print(f'  列 {col_letter}: 宽度 = {width}')

# 行高
print('\n行高信息:')
for row_idx in range(1, ws.max_row + 1):
    height = ws.row_dimensions[row_idx].height
    if height:
        print(f'  行 {row_idx}: 高度 = {height}')

# 详细分析每个单元格的格式
print('\n单元格格式详细分析:')
print('-' * 80)

for row_idx in range(1, min(ws.max_row + 1, 10)):
    for col_idx in range(1, min(ws.max_column + 1, 15)):
        cell = ws.cell(row=row_idx, column=col_idx)
        
        # 跳过空单元格（如果有）
        if cell.value is None and not cell.font.bold and cell.fill.fgColor.index == '00000000':
            continue
            
        cell_letter = f'{get_column_letter(col_idx)}{row_idx}'
        print(f'\n单元格 {cell_letter}:')
        print(f'  值: {cell.value}')
        
        # 字体
        font = cell.font
        print(f'  字体: {font.name}, 大小: {font.size}, 加粗: {font.bold}, 斜体: {font.italic}')
        
        # 对齐方式
        alignment = cell.alignment
        print(f'  对齐: 水平={alignment.horizontal}, 垂直={alignment.vertical}, 换行={alignment.wrap_text}')
        
        # 背景颜色
        fill = cell.fill
        if fill.fgColor.index != '00000000':
            print(f'  背景颜色: {fill.fgColor.index}')
        
        # 边框
        border = cell.border
        has_border = False
        border_info = []
        if border.left.style:
            border_info.append(f'左: {border.left.style}')
            has_border = True
        if border.right.style:
            border_info.append(f'右: {border.right.style}')
            has_border = True
        if border.top.style:
            border_info.append(f'上: {border.top.style}')
            has_border = True
        if border.bottom.style:
            border_info.append(f'下: {border.bottom.style}')
            has_border = True
        if has_border:
            print(f'  边框: {", ".join(border_info)}')

print('\n' + '=' * 100)
print('分析完成！')