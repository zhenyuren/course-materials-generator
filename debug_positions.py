from openpyxl import load_workbook

# 读取生成的文件
file_path = '智能金融学院_未知教师/FIN3004A_大数据分析基础/大数据分析基础-课程组期初教学资料检查情况记录表-任渝.xlsx'
wb = load_workbook(file_path)
ws = wb.active

print('调试单元格位置:')
print('=' * 80)

# 检查第20行和第21行的单元格
print('\n第20行单元格值:')
for col in range(1, 15):
    cell = ws.cell(row=20, column=col)
    if cell.value is not None:
        print(f'第20行第{col}列: {cell.value}')

print('\n第21行单元格值:')
for col in range(1, 15):
    cell = ws.cell(row=21, column=col)
    if cell.value is not None:
        print(f'第21行第{col}列: {cell.value}')

print('\n第19行单元格值:')
for col in range(1, 15):
    cell = ws.cell(row=19, column=col)
    if cell.value is not None:
        print(f'第19行第{col}列: {cell.value}')