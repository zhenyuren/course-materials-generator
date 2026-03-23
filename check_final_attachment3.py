from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# 查看生成的附件3文件
wb = load_workbook('智能金融学院_未知教师/FIN3004A_大数据分析基础/大数据分析基础-课程组期初教学资料检查情况记录表-任渝.xlsx')
ws = wb.active

print('生成的附件3文件详细分析:')
print('=' * 100)

# 基本信息
print(f'工作表名称: {ws.title}')
print(f'总行数: {ws.max_row}')
print(f'总列数: {ws.max_column}')

# 合并单元格
merged_list = list(ws.merged_cells.ranges)
print(f'\n合并单元格数量: {len(merged_list)}')
for i, merged_cell in enumerate(merged_list, 1):
    print(f'  {i}. {merged_cell}')

# 列宽
print('\n列宽信息:')
for col_idx in range(1, ws.max_column + 1):
    col_letter = get_column_letter(col_idx)
    width = ws.column_dimensions[col_letter].width
    print(f'  列 {col_letter}: 宽度 = {width}')

# 行高
print('\n行高信息:')
for row_idx in range(1, 6):
    height = ws.row_dimensions[row_idx].height
    if height:
        print(f'  行 {row_idx}: 高度 = {height}')

# 详细分析关键单元格的格式
print('\n关键单元格格式详细分析:')
print('-' * 80)

# 标题行
cell = ws.cell(row=1, column=1)
print(f'\n标题单元格 A1:')
print(f'  值: {cell.value}')
print(f'  字体: {cell.font.name}, 大小: {cell.font.size}, 加粗: {cell.font.bold}')
print(f'  对齐: 水平={cell.alignment.horizontal}, 垂直={cell.alignment.vertical}')

# 表头行
print('\n表头行:')
headers = ['序号', '课程代码', '课程名', '开课单位', '使用年级/层次/专业', '归属校区', '教师（执笔人）', '课程组验收负责人', '教案评价指标']
for i, header in enumerate(headers, 1):
    cell = ws.cell(row=2, column=i)
    print(f'  单元格 {get_column_letter(i)}2: 值="{cell.value}", 字体={cell.font.name}, 大小={cell.font.size}, 加粗={cell.font.bold}')

# 评价指标行
print('\n评价指标行:')
metrics = ['教学大纲符合度', '教案撰写规范性', '教学方式', '教学进度及安排', '课程设计和框架', '总分']
for i, metric in enumerate(metrics, 9):
    cell = ws.cell(row=3, column=i)
    print(f'  单元格 {get_column_letter(i)}3: 值="{cell.value}", 字体={cell.font.name}, 大小={cell.font.size}')

# 分数行
print('\n分数行:')
scores = ['15分', '10分', '10分', '15分', '50分', '100分']
for i, score in enumerate(scores, 9):
    cell = ws.cell(row=4, column=i)
    print(f'  单元格 {get_column_letter(i)}4: 值="{cell.value}", 字体={cell.font.name}, 大小={cell.font.size}')

# 数据行
print('\n数据行:')
data_cells = [
    ('课程代码', 2),
    ('课程名', 3),
    ('开课单位', 4),
    ('使用年级/层次/专业', 6),
    ('教师（执笔人）', 7),
    ('课程组验收负责人', 8)
]
for label, col_idx in data_cells:
    cell = ws.cell(row=5, column=col_idx)
    print(f'  {label}: {cell.value}')

# 检查边框
print('\n边框检查:')
border_cells = [
    ('A1', 1, 1),
    ('B2', 2, 2),
    ('C3', 3, 3),
    ('D4', 4, 4),
    ('E5', 5, 5)
]
for label, row_idx, col_idx in border_cells:
    cell = ws.cell(row=row_idx, column=col_idx)
    border = cell.border
    has_border = any([border.left.style, border.right.style, border.top.style, border.bottom.style])
    print(f'  单元格 {label}: 有边框 = {has_border}')

print('\n' + '=' * 100)
print('分析完成！')