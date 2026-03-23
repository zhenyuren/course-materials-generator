from openpyxl import load_workbook

# 验证财经公文写作课程的数据填充
file_path = "智能金融学院_未知教师/FIN0008B_财经公文写作/财经公文写作-教学大纲审核汇总表-未知教师.xlsx"
wb = load_workbook(file_path)
ws = wb.active

print("财经公文写作课程数据填充验证：")
print(f"课程代码: {ws.cell(row=5, column=2).value}")
print(f"课程名称: {ws.cell(row=5, column=3).value}")
print(f"开课单位: {ws.cell(row=5, column=4).value}")
print(f"所属校区: {ws.cell(row=5, column=5).value}")
print(f"使用年级/层次/专业: {ws.cell(row=5, column=6).value}")
print(f"执笔人: {ws.cell(row=5, column=7).value}")
print(f"验收负责人: {ws.cell(row=5, column=8).value}")

print("\n大数据分析基础课程数据填充验证：")
file_path2 = "智能金融学院_未知教师/FIN3004A_大数据分析基础/大数据分析基础-教学大纲审核汇总表-未知教师.xlsx"
wb2 = load_workbook(file_path2)
ws2 = wb2.active

print(f"课程代码: {ws2.cell(row=5, column=2).value}")
print(f"课程名称: {ws2.cell(row=5, column=3).value}")
print(f"开课单位: {ws2.cell(row=5, column=4).value}")
print(f"所属校区: {ws2.cell(row=5, column=5).value}")
print(f"使用年级/层次/专业: {ws2.cell(row=5, column=6).value}")
print(f"执笔人: {ws2.cell(row=5, column=7).value}")
print(f"验收负责人: {ws2.cell(row=5, column=8).value}")
