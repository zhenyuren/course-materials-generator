from openpyxl import load_workbook

# 查看原始模板文件
wb = load_workbook('期初资料/附件3 课程组教学资料检查情况记录表_原始.xlsx')
ws = wb.active

print('原始附件3模板文件内容:')
print('=' * 80)

# 打印标题行
title = ws.cell(row=1, column=1).value
print(f'标题行: {title}')

# 打印表头和数据行
print('\n文件内容预览:')
for row_idx in range(1, 20):
    row_values = []
    for col_idx in range(1, 15):
        cell_value = ws.cell(row=row_idx, column=col_idx).value
        row_values.append(str(cell_value) if cell_value is not None else '')
    print(f'第{row_idx}行: ' + ' | '.join(row_values))

# 检查合并单元格
print('\n合并单元格:')
merged_list = list(ws.merged_cells.ranges)
for merged_cell in merged_list:
    print(f'  {merged_cell}')

# 检查是否有占位符
print('\n搜索占位符:')
for row_idx in range(1, ws.max_row + 1):
    for col_idx in range(1, ws.max_column + 1):
        cell_value = ws.cell(row=row_idx, column=col_idx).value
        if cell_value and ('{{' in str(cell_value) or '}}' in str(cell_value)):
            print(f'第{row_idx}行第{col_idx}列: {cell_value}')