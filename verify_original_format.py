from openpyxl import load_workbook

# 读取原始模板
wb_template = load_workbook('期初资料/附件3 课程组教学资料检查情况记录表-原始格式.xlsx')
ws_template = wb_template.active

# 读取生成的文件
wb_generated = load_workbook('智能金融学院_未知教师/FIN3004A_大数据分析基础/大数据分析基础-课程组期初教学资料检查情况记录表-任渝.xlsx')
ws_generated = wb_generated.active

print('原始格式验证:')
print('=' * 80)

# 打印原始模板和生成文件的内容对比
print('\n原始模板内容:')
for row_idx in range(1, 25):
    row_values = []
    for col_idx in range(1, 15):
        cell_value = ws_template.cell(row=row_idx, column=col_idx).value
        row_values.append(str(cell_value) if cell_value is not None else '')
    print(f'第{row_idx}行: ' + ' | '.join(row_values))

print('\n' + '=' * 80)
print('生成文件内容:')
for row_idx in range(1, 25):
    row_values = []
    for col_idx in range(1, 15):
        cell_value = ws_generated.cell(row=row_idx, column=col_idx).value
        row_values.append(str(cell_value) if cell_value is not None else '')
    print(f'第{row_idx}行: ' + ' | '.join(row_values))

# 检查合并单元格
print('\n' + '=' * 80)
print('合并单元格对比:')

print('原始模板合并单元格:')
merged_template = list(ws_template.merged_cells.ranges)
for merged in merged_template:
    print(f'  {merged}')

print('\n生成文件合并单元格:')
merged_generated = list(ws_generated.merged_cells.ranges)
for merged in merged_generated:
    print(f'  {merged}')

# 检查关键格式
print('\n' + '=' * 80)
print('关键格式检查:')

# 检查评价指标列背景色
print('\n评价指标列背景色:')
for col_idx in range(9, 15):
    cell_template = ws_template.cell(row=3, column=col_idx)
    cell_generated = ws_generated.cell(row=3, column=col_idx)
    print(f'列{col_idx}: 模板={cell_template.fill.fgColor.index}, 生成={cell_generated.fill.fgColor.index}')

# 检查分数列背景色
print('\n分数列背景色:')
for col_idx in range(9, 15):
    cell_template = ws_template.cell(row=4, column=col_idx)
    cell_generated = ws_generated.cell(row=4, column=col_idx)
    print(f'列{col_idx}: 模板={cell_template.fill.fgColor.index}, 生成={cell_generated.fill.fgColor.index}')

# 检查表末尾端结构
print('\n表末尾端结构:')
for row_idx in [19, 20, 21, 22, 23, 24]:
    template_value = ws_template.cell(row=row_idx, column=1).value
    generated_value = ws_generated.cell(row=row_idx, column=1).value
    print(f'第{row_idx}行: 模板="{template_value}", 生成="{generated_value}"')