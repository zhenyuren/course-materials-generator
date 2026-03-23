from openpyxl import load_workbook

def analyze_excel_format(file_path, title):
    """分析Excel文件的格式信息"""
    print(f"\n{'='*60}")
    print(f"{title}")
    print(f"{'='*60}")
    
    wb = load_workbook(file_path)
    ws = wb.active
    
    print(f"文件名: {file_path}")
    print(f"工作表: {ws.title}")
    print(f"总行数: {ws.max_row}")
    print(f"总列数: {ws.max_column}")
    
    # 检查合并单元格
    merged_cells = ws.merged_cells.ranges
    print(f"\n合并单元格数量: {len(merged_cells)}")
    merged_list = list(merged_cells)
    for i, merged_cell in enumerate(merged_list[:5], 1):
        print(f"  {i}. {merged_cell}")
    if len(merged_list) > 5:
        print(f"  ... 还有 {len(merged_list) - 5} 个合并单元格")
    
    # 检查列宽
    print(f"\n列宽信息:")
    for col_idx in range(1, min(ws.max_column + 1, 10)):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        width = ws.column_dimensions[col_letter].width
        print(f"  列 {col_letter}: {width}")
    
    # 检查行高
    print(f"\n行高信息:")
    for row_idx in range(1, min(ws.max_row + 1, 10)):
        height = ws.row_dimensions[row_idx].height
        print(f"  行 {row_idx}: {height}")
    
    # 检查前5行内容
    print(f"\n前5行内容:")
    for row_idx in range(1, min(ws.max_row + 1, 6)):
        row_values = []
        for col_idx in range(1, min(ws.max_column + 1, 10)):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            row_values.append(str(cell_value) if cell_value is not None else 'None')
        print(f"  行 {row_idx}: {', '.join(row_values)}")
    
    return wb

# 分析附件3模板和生成文件
template3 = "期初资料/附件3 课程组教学资料检查情况记录表_new.xlsx"
generated3 = "智能金融学院_未知教师/FIN3004A_大数据分析基础/大数据分析基础-课程组期初教学资料检查情况记录表-任渝.xlsx"

analyze_excel_format(template3, "附件3模板文件分析")
analyze_excel_format(generated3, "附件3生成文件分析")

# 分析附件1模板和生成文件
template1 = "期初资料/附件1 教学大纲的审核汇总表-大数据信息分析_任渝_new.xlsx"
generated1 = "智能金融学院_未知教师/FIN3004A_大数据分析基础/数字经济-教学大纲审核汇总表-任渝.xlsx"

analyze_excel_format(template1, "附件1模板文件分析")
analyze_excel_format(generated1, "附件1生成文件分析")