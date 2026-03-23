from openpyxl import load_workbook

# 查看生成的附件3文件结构
wb = load_workbook('智能金融学院_未知教师/FIN3004A_大数据分析基础/大数据分析基础-课程组期初教学资料检查情况记录表-任渝.xlsx')
ws = wb.active
print('附件3生成文件结构:')
print(f'总行数: {ws.max_row}')
print(f'总列数: {ws.max_column}')
print('前8行内容:')
for row in ws.iter_rows(min_row=1, max_row=8, values_only=True):
    print(row)

print('\n' + '='*50 + '\n')

# 查看生成的附件1文件结构
wb = load_workbook('智能金融学院_未知教师/FIN3004A_大数据分析基础/数字经济-教学大纲审核汇总表-任渝.xlsx')
ws = wb.active
print('附件1生成文件结构:')
print(f'总行数: {ws.max_row}')
print(f'总列数: {ws.max_column}')
print('前8行内容:')
for row in ws.iter_rows(min_row=1, max_row=8, values_only=True):
    print(row)