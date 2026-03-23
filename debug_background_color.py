from openpyxl import load_workbook

# 查看生成的附件3文件
wb = load_workbook('智能金融学院_未知教师/FIN3004A_大数据分析基础/大数据分析基础-课程组期初教学资料检查情况记录表-任渝.xlsx')
ws = wb.active

print('背景色调试信息:')
print('=' * 80)

# 检查评价指标列的背景色
print('\n评价指标列背景色:')
for col_idx in range(9, 15):
    cell = ws.cell(row=3, column=col_idx)
    fill = cell.fill
    print(f'单元格 I{col_idx}:')
    print(f'  fgColor.index: {fill.fgColor.index}')
    print(f'  bgColor.index: {fill.bgColor.index}')
    print(f'  fill_type: {fill.fill_type}')

# 检查分数列的背景色
print('\n分数列背景色:')
for col_idx in range(9, 15):
    cell = ws.cell(row=4, column=col_idx)
    fill = cell.fill
    print(f'单元格 I{col_idx}:')
    print(f'  fgColor.index: {fill.fgColor.index}')
    print(f'  bgColor.index: {fill.bgColor.index}')
    print(f'  fill_type: {fill.fill_type}')

# 直接检查模板文件
print('\n' + '=' * 80)
print('检查模板文件:')
wb_template = load_workbook('期初资料/附件3 课程组教学资料检查情况记录表-完整版模板.xlsx')
ws_template = wb_template.active

print('\n模板文件评价指标列背景色:')
for col_idx in range(9, 15):
    cell = ws_template.cell(row=3, column=col_idx)
    fill = cell.fill
    print(f'单元格 I{col_idx}:')
    print(f'  fgColor.index: {fill.fgColor.index}')
    print(f'  bgColor.index: {fill.bgColor.index}')
    print(f'  fill_type: {fill.fill_type}')

print('\n模板文件分数列背景色:')
for col_idx in range(9, 15):
    cell = ws_template.cell(row=4, column=col_idx)
    fill = cell.fill
    print(f'单元格 I{col_idx}:')
    print(f'  fgColor.index: {fill.fgColor.index}')
    print(f'  bgColor.index: {fill.bgColor.index}')
    print(f'  fill_type: {fill.fill_type}')